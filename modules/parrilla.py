# modules/parrilla.py — Parrilla de Contenido generada con IA

import os
import json
import base64
import calendar as cal_mod
from datetime import date
from pathlib import Path
import io
import pandas as pd
import streamlit as st

ROOT = Path(__file__).parent.parent

_DIA_MAP = {
    'lunes': 0, 'martes': 1, 'miércoles': 2, 'miercoles': 2,
    'jueves': 3, 'viernes': 4, 'sábado': 5, 'sabado': 5, 'domingo': 6,
}
_DIA_LABEL = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

MESES_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}

_MASTER_PROMPTS = {
    'isometric': (
        "PREMIUM ISOMETRIC SMART-CITY DIORAMA. Ultra-clean corporate technology visualization. "
        "Photorealistic miniature world built as a highly detailed architectural scale model. "
        "Floating rounded platform with premium chamfered edges and integrated blue accent lighting. "
        "Product-render presentation aesthetic. High-end executive showcase quality. "
        "Apple keynote product visualization style. Luxury industrial design language. "
        "Architectural model presentation. Realistic urban planning geometry. "
        "Highly detailed roads, sidewalks, landscaping, infrastructure, buildings and public spaces. "
        "Technology naturally integrated into the environment. Infrastructure-first design philosophy. "
        "Connected systems represented through physical infrastructure and architectural elements. "
        "Elegant blue, cyan and white lighting accents. Subtle integrated illumination throughout the environment. "
        "Premium smart-city aesthetic. Clean and organized composition. Balanced visual hierarchy. "
        "Generous negative space around the platform. "
        "Realistic scale and proportions. Photorealistic vehicles, vegetation, buildings and infrastructure. "
        "Physically accurate materials. PBR materials. Ray-traced reflections. Global illumination. "
        "Soft ambient occlusion. Photorealistic shadows. Ultra-detailed surface textures. "
        "Octane Render quality. Unreal Engine 5 quality. Cinema 4D product-render quality. 8K. "
        "Hyperrealistic. Miniature-world realism. Corporate technology showcase aesthetic. "
        "NO cyberpunk. NO sci-fi interfaces. NO holograms. NO floating dashboards. "
        "NO augmented reality overlays. NO futuristic fantasy architecture. NO excessive glow. "
        "NO visual clutter. NO visual noise. NO text. NO logos. NO watermarks."
    ),
    'cards': (
        "PREMIUM CORPORATE SECURITY TECHNOLOGY VISUALIZATION. Ultra-clean enterprise marketing aesthetic. "
        "Photorealistic commercial advertising style. Real-world environments combined with premium 3D technology elements. "
        "White and light gray minimalist composition. Large negative space. Executive presentation quality. "
        "Hero technology device prominently displayed. Elegant blue technology accents. "
        "Glowing blue circuit pathways connecting devices, infrastructure, vehicles and people. "
        "Premium floating information panels. Minimalist UI cards. Clean information hierarchy. "
        "Luxury technology marketing design. Apple enterprise presentation aesthetic. "
        "Verkada, Motorola Solutions and Genetec visual language. Professional corporate advertising composition. "
        "Photorealistic vehicles, people, buildings and infrastructure. Physically accurate materials. "
        "PBR rendering. Ray-traced reflections. Global illumination. Soft ambient occlusion. "
        "Photorealistic shadows. Realistic reflections on pavement, glass and vehicles. Ultra detailed. "
        "Commercial photography quality. Octane Render quality. Unreal Engine 5 quality. "
        "Cinema 4D product-render quality. 8K. Hyperrealistic. Blue, white, gray and black color palette. "
        "NO cyberpunk. NO holograms. NO sci-fi interfaces. NO floating dashboards. "
        "NO augmented reality overlays. NO fantasy technology. NO excessive glow. "
        "NO visual clutter. NO gaming aesthetic. NO visual noise. NO logos. NO watermarks."
    ),
    'infographic': (
        "PREMIUM SMART CITY DIGITAL TWIN VISUALIZATION. Ultra-clean executive technology presentation. "
        "Hyperrealistic city environment blended with elegant operational data visualization. "
        "White and light gray corporate aesthetic. Clean composition with large negative space. "
        "Glassmorphism UI cards with subtle transparency and soft blur. "
        "Floating information panels integrated naturally into the scene. "
        "Premium command-and-control visual language. Connected infrastructure represented through "
        "glowing blue pathways, location markers and operational flows. "
        "Realistic vehicles, emergency units, operators, cameras, sensors and urban infrastructure. "
        "High-end public safety and smart-city branding aesthetic. Photorealistic city environments. "
        "Architectural realism. Soft atmospheric depth. Elegant blue accent lighting. "
        "Executive infographic quality. Apple enterprise presentation aesthetic. Digital twin city visualization. "
        "Physically accurate materials. PBR rendering. Global illumination. Ray-traced reflections. "
        "Soft shadows. Commercial photography quality. Octane Render quality. Unreal Engine 5 quality. "
        "Cinema 4D product-render quality. Ultra detailed. 8K. Hyperrealistic. "
        "NO cyberpunk. NO holograms. NO sci-fi interfaces. NO floating dashboards. "
        "NO fantasy technology. NO excessive glow. NO visual clutter. NO gaming aesthetic. "
        "NO visual noise. NO dark military aesthetic. NO logos. NO watermarks."
    ),
}

_STYLE_OPTIONS = {
    '🏙️ Isométrico':  'isometric',
    '🃏 Cards':        'cards',
    '📊 Infografía':   'infographic',
    '✏️ Libre':        'libre',
}

OBJETIVOS_PRESET = [
    "Posicionamiento y liderazgo en el sector seguridad pública",
    "Generación de interés y leads en nuevos territorios",
    "Educación técnica: mostrar capacidades de la plataforma",
    "Construcción de confianza y prueba social",
    "Presencia y cobertura de evento o temporada relevante",
    "Lanzamiento o anuncio de nuevo producto / módulo",
    "Personalizado…",
]

_ESTADOS = ['Borrador', 'En diseño', 'Aprobado', 'Publicado']

_ESTADO_COLORS = {
    'Borrador':  {
        'bg':     '#1c1f2e',
        'border': '#6b7280',
        'text':   '#d1d5db',
        'badge':  '#4b5563',
        'solid':  '#6b7280',   # for filled indicators
    },
    'En diseño': {
        'bg':     '#2d1f02',
        'border': '#f59e0b',
        'text':   '#fde68a',
        'badge':  '#b45309',
        'solid':  '#f59e0b',
    },
    'Aprobado':  {
        'bg':     '#0a1e4a',
        'border': '#60a5fa',
        'text':   '#bfdbfe',
        'badge':  '#2563eb',
        'solid':  '#3b82f6',
    },
    'Publicado': {
        'bg':     '#052e16',
        'border': '#22c55e',
        'text':   '#bbf7d0',
        'badge':  '#16a34a',
        'solid':  '#22c55e',
    },
}

_ESTADO_EMOJI = {
    'Borrador':  '⚪',
    'En diseño': '🟡',
    'Aprobado':  '🔵',
    'Publicado': '🟢',
}

# Display labels used in the data_editor selectbox (emoji prefix for visual color cue)
_ESTADOS_DISPLAY   = [f"{_ESTADO_EMOJI[e]} {e}" for e in _ESTADOS]
_DISPLAY_TO_ESTADO = {f"{_ESTADO_EMOJI[e]} {e}": e for e in _ESTADOS}

# Allow color lookups with either clean name or display name
for _e in list(_ESTADOS):
    _ESTADO_COLORS[f"{_ESTADO_EMOJI[_e]} {_e}"] = _ESTADO_COLORS[_e]


def _to_display(estado):
    emoji = _ESTADO_EMOJI.get(estado, '')
    return f"{emoji} {estado}" if emoji else estado


def _from_display(display):
    return _DISPLAY_TO_ESTADO.get(display, display)


# ── Env / Brand ────────────────────────────────────────────────────────────────

def _load_env():
    env_path = ROOT / '.env'
    if env_path.exists():
        for line in env_path.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, _, val = line.partition('=')
                os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))
    try:
        import streamlit as st
        for key, val in st.secrets.items():
            if isinstance(val, str):
                os.environ.setdefault(key, val)
    except Exception:
        pass


def _load_brand(brand_id):
    p = ROOT / 'marca' / brand_id / 'brand.json'
    return json.loads(p.read_text(encoding='utf-8')) if p.exists() else {}


def _get_api_key():
    _load_env()
    return os.environ.get('ANTHROPIC_API_KEY', '').strip()


def _get_gemini_key():
    _load_env()
    return os.environ.get('GEMINI_API_KEY', '').strip()


def _call_gemini_raw(prompt):
    """Calls Gemini generateContent and returns the raw text response."""
    import urllib.request
    key = _get_gemini_key()
    url = f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={key}'
    body = json.dumps({
        'contents': [{'parts': [{'text': prompt}]}],
        'generationConfig': {
            'maxOutputTokens': 65536,
            'temperature': 0.7,
        },
        'safetySettings': [
            {'category': 'HARM_CATEGORY_HARASSMENT',       'threshold': 'BLOCK_NONE'},
            {'category': 'HARM_CATEGORY_HATE_SPEECH',      'threshold': 'BLOCK_NONE'},
            {'category': 'HARM_CATEGORY_SEXUALLY_EXPLICIT','threshold': 'BLOCK_NONE'},
            {'category': 'HARM_CATEGORY_DANGEROUS_CONTENT','threshold': 'BLOCK_NONE'},
        ],
    }).encode()
    req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
    resp = urllib.request.urlopen(req, timeout=180)
    data = json.loads(resp.read())
    candidates = data.get('candidates', [])
    if not candidates:
        prompt_fb = data.get('promptFeedback', {})
        raise RuntimeError(f"Gemini bloqueó la solicitud: {prompt_fb.get('blockReason', 'razón desconocida')}")
    cand   = candidates[0]
    reason = cand.get('finishReason', '')
    if reason == 'SAFETY':
        raise RuntimeError("Gemini bloqueó la respuesta por filtros de seguridad. Intenta con Claude.")
    text = ''
    for part in cand.get('content', {}).get('parts', []):
        text += part.get('text', '')
    if not text.strip():
        raise RuntimeError(f"Gemini devolvió respuesta vacía (finishReason: {reason}).")
    return text


def _stream_gemini(prompt):
    """Generator: yields text chunks from Gemini streaming API."""
    import urllib.request
    key = _get_gemini_key()
    url = (f'https://generativelanguage.googleapis.com/v1beta/models/'
           f'gemini-2.5-flash:streamGenerateContent?alt=sse&key={key}')
    body = json.dumps({
        'contents': [{'parts': [{'text': prompt}]}],
        'generationConfig': {'maxOutputTokens': 8192},
    }).encode()
    req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
    resp = urllib.request.urlopen(req, timeout=120)
    for line in resp:
        line = line.decode('utf-8').strip()
        if line.startswith('data: '):
            data_str = line[6:]
            if data_str == '[DONE]':
                break
            try:
                data = json.loads(data_str)
                for cand in data.get('candidates', []):
                    for part in cand.get('content', {}).get('parts', []):
                        text = part.get('text', '')
                        if text:
                            yield text
            except Exception:
                pass


def _ai_provider():
    try:
        return st.session_state.get('ai_provider', 'gemini')
    except Exception:
        return 'gemini'


def _composite_logo(img_bytes, brand, position='bottom-right', logo_pct=0.20,
                    logo_variant='white', custom_logo_bytes=None):
    """Overlay the brand logo on the generated image. Returns PNG bytes."""
    from PIL import Image

    if custom_logo_bytes:
        try:
            logo_img = Image.open(io.BytesIO(custom_logo_bytes)).convert('RGBA')
        except Exception:
            return img_bytes
    else:
        brand_id  = brand.get('brand_id', '')
        logo_path = ROOT / 'marca' / brand_id / 'logos' / f'{logo_variant}.png'
        if not logo_path.exists():
            logo_path = ROOT / 'marca' / brand_id / 'logos' / 'white.png'
        if not logo_path.exists():
            return img_bytes
        logo_img = Image.open(str(logo_path)).convert('RGBA')

    base_img = Image.open(io.BytesIO(img_bytes)).convert('RGBA')

    bw, bh   = base_img.size
    logo_w   = int(bw * logo_pct)
    ratio    = logo_w / logo_img.width
    logo_h   = int(logo_img.height * ratio)
    logo_img = logo_img.resize((logo_w, logo_h), Image.LANCZOS)

    pad = int(bw * 0.03)
    if position == 'bottom-right':
        x = bw - logo_w - pad; y = bh - logo_h - pad
    elif position == 'bottom-left':
        x, y = pad, bh - logo_h - pad
    elif position == 'top-right':
        x, y = bw - logo_w - pad, pad
    elif position == 'center':
        x, y = (bw - logo_w) // 2, (bh - logo_h) // 2
    else:
        x, y = pad, pad

    base_img.paste(logo_img, (x, y), logo_img)

    out = io.BytesIO()
    base_img.convert('RGB').save(out, format='PNG')
    return out.getvalue()


def _logo_adder_ui(key_prefix, img_bytes, brand):
    """
    Inline logo overlay widget. Returns composited PNG bytes when Apply is clicked, else None.
    _open_key is intentionally NOT used as a widget key= parameter so it can be
    set freely at any time without triggering Streamlit's
    'cannot be modified after widget is instantiated' error.
    """
    # _la_panel_ prefix — intentionally different from the old _la_open_ widget key
    # so that any lingering browser session state from the previous st.toggle
    # implementation does not interfere with setting this plain variable.
    _panel_key = f"_la_panel_{key_prefix}"
    _cust_key  = f"_la_cust_{key_prefix}"

    _is_open = st.session_state.get(_panel_key, False)

    # Plain button for open/close — _panel_key is never used as a widget key=,
    # so Streamlit never "owns" it and it can be set freely at any point.
    _btn_lbl = "✕ Cerrar logo" if _is_open else "🏷️ Agregar logo"
    if st.button(_btn_lbl, key=f"_la_btn_{key_prefix}"):
        _is_open = not _is_open
        st.session_state[_panel_key] = _is_open

    if not _is_open:
        return None

    _brand_id   = brand.get('brand_id', '')
    _logo_dir   = ROOT / 'marca' / _brand_id / 'logos'
    _avail_vars = [v for v in ('white', 'color') if (_logo_dir / f'{v}.png').exists()]
    _var_labels = {'white': '⬜ Blanco', 'color': '🎨 Color'}
    _has_custom = bool(st.session_state.get(_cust_key))

    # Show logo thumbnails
    if _avail_vars:
        _thumb_cols = st.columns(len(_avail_vars))
        for _vi, _vv in enumerate(_avail_vars):
            _thumb_cols[_vi].image(str(_logo_dir / f'{_vv}.png'), use_container_width=True)

        if not _has_custom:
            st.radio(
                "Variante",
                _avail_vars,
                format_func=lambda v: _var_labels.get(v, v),
                key=f"_la_var_{key_prefix}",
                horizontal=True,
                label_visibility="collapsed",
            )

    # Custom logo uploader
    _cust_up = st.file_uploader(
        "O sube tu propio logo (PNG/JPG)",
        type=['png', 'jpg', 'jpeg'],
        key=f"_la_uploader_{key_prefix}",
    )
    if _cust_up is not None:
        _raw = _cust_up.read()
        if _raw:
            from PIL import Image as _PIL_la
            _cl = _PIL_la.open(io.BytesIO(_raw)).convert('RGBA')
            _cb = io.BytesIO(); _cl.save(_cb, format='PNG')
            st.session_state[_cust_key] = _cb.getvalue()
            _has_custom = True

    if _has_custom:
        _hc1, _hc2 = st.columns([5, 1])
        _hc1.caption("✅ Logo propio activo")
        with _hc2:
            if st.button("🗑️", key=f"_la_rmcust_{key_prefix}", help="Quitar logo propio"):
                st.session_state.pop(_cust_key, None)

    # Position + size in two columns
    _pos_opts = {
        '↘ Inf. derecha':   'bottom-right',
        '↙ Inf. izquierda': 'bottom-left',
        '↗ Sup. derecha':   'top-right',
        '↖ Sup. izquierda': 'top-left',
        '🔲 Centro':         'center',
    }
    _col_pos, _col_sz = st.columns([3, 2])
    with _col_pos:
        _sel_pos = st.selectbox(
            "Posición",
            list(_pos_opts.keys()),
            key=f"_la_pos_sel_{key_prefix}",
        )
    with _col_sz:
        _pct_val = st.slider(
            "Tamaño (%)", min_value=8, max_value=40,
            value=20, step=2,
            key=f"_la_pct_sl_{key_prefix}",
        )

    # Live preview — updates automatically on every control change
    _prev_var  = st.session_state.get(f"_la_var_{key_prefix}", _avail_vars[0] if _avail_vars else 'white')
    if _has_custom:
        _prev_var = 'white'
    _prev_pos  = _pos_opts.get(_sel_pos, 'bottom-right')
    _prev_pct  = _pct_val / 100
    _prev_cust = st.session_state.get(_cust_key) if _has_custom else None
    try:
        _preview_bytes = _composite_logo(
            img_bytes, brand,
            position=_prev_pos,
            logo_pct=_prev_pct,
            logo_variant=_prev_var,
            custom_logo_bytes=_prev_cust,
        )
        st.caption("Vista previa:")
        st.image(_preview_bytes, use_container_width=True)
    except Exception:
        pass

    if st.button("✅ Aplicar logo", key=f"_la_apply_{key_prefix}", type="primary", use_container_width=True):
        try:
            _result = _composite_logo(
                img_bytes, brand,
                position=_prev_pos,
                logo_pct=_prev_pct,
                logo_variant=_prev_var,
                custom_logo_bytes=_prev_cust,
            )
            st.session_state[_panel_key] = False
            return _result
        except Exception as _le:
            st.error(f"Error al aplicar logo: {_le}")

    return None


def _edit_imagen_gemini(image_bytes, instruction):
    """Send image + text instruction to Gemini image editing model. Returns PNG bytes."""
    import urllib.request
    key = _get_gemini_key()
    img_b64 = base64.b64encode(image_bytes).decode()
    url = (f'https://generativelanguage.googleapis.com/v1beta/models/'
           f'gemini-2.5-flash-image:generateContent?key={key}')
    _system = (
        "Eres un editor de imágenes profesional. "
        "Aplica exactamente la instrucción del usuario a la imagen. "
        "REGLA OBLIGATORIA: cualquier texto visible en la imagen resultante "
        "debe estar escrito en español. Si la instrucción pide añadir, cambiar "
        "o reemplazar texto, escríbelo en español. "
        "Puedes editar textos, colores, fondos, composición y elementos visuales."
    )
    body = json.dumps({
        'system_instruction': {'parts': [{'text': _system}]},
        'contents': [{
            'parts': [
                {'inlineData': {'mimeType': 'image/png', 'data': img_b64}},
                {'text': instruction},
            ]
        }],
        'generationConfig': {'responseModalities': ['IMAGE', 'TEXT']},
    }).encode()
    req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
    resp = urllib.request.urlopen(req, timeout=180)
    data = json.loads(resp.read())
    parts = data.get('candidates', [{}])[0].get('content', {}).get('parts', [])
    for p in parts:
        if 'inlineData' in p:
            raw = base64.b64decode(p['inlineData']['data'])
            # Normalize to PNG
            from PIL import Image
            import io
            img = Image.open(io.BytesIO(raw)).convert('RGB')
            buf = io.BytesIO()
            img.save(buf, format='PNG')
            return buf.getvalue()
    raise RuntimeError('El modelo no devolvió imagen editada. Intenta con otra instrucción.')


_SPANISH_TEXT_SUFFIX = (
    " Any text visible in the image must be written in Spanish / "
    "Cualquier texto visible en la imagen debe estar en español."
)

# Font search paths — tried in order, first match wins
_FONT_PATHS = [
    '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
    '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    '/usr/share/fonts/truetype/ubuntu/Ubuntu-B.ttf',
    '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
    '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf',
    '/System/Library/Fonts/Helvetica.ttc',
    '/System/Library/Fonts/SFNS.ttf',
    'C:/Windows/Fonts/arialbd.ttf',
    'C:/Windows/Fonts/arial.ttf',
]


def _load_font(size, brand_id=None):
    """Return the best available PIL font at the given size.
    Checks brand-specific fonts first (marca/<brand_id>/fonts/)."""
    from PIL import ImageFont
    paths = []
    if brand_id:
        brand_font_dir = ROOT / 'marca' / brand_id / 'fonts'
        for candidate in ('SpaceGrotesk-Bold.ttf', 'bold.ttf', 'regular.ttf'):
            paths.append(str(brand_font_dir / candidate))
    paths += _FONT_PATHS
    for fp in paths:
        try:
            return ImageFont.truetype(fp, size)
        except Exception:
            pass
    try:
        return ImageFont.load_default(size=size)
    except TypeError:
        return ImageFont.load_default()


def _hex_to_color_name(hex_str):
    """Convert a hex color string to a descriptive Spanish name so AI models
    don't render the hex code as literal text in the image."""
    try:
        h = hex_str.strip('#')
        if len(h) == 3:
            h = h[0]*2 + h[1]*2 + h[2]*2
        if len(h) != 6:
            return hex_str
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        mx, mn = max(r, g, b), min(r, g, b)
        brightness = mx / 255
        saturation = (mx - mn) / mx if mx > 0 else 0

        if brightness < 0.18:
            return 'negro profundo'
        if brightness > 0.88 and saturation < 0.12:
            return 'blanco'
        if saturation < 0.12:
            if brightness < 0.35:
                return 'gris muy oscuro'
            if brightness < 0.6:
                return 'gris medio'
            return 'gris claro'

        d = mx - mn
        if mx == r:
            h_deg = (60 * ((g - b) / d)) % 360
        elif mx == g:
            h_deg = 60 * ((b - r) / d) + 120
        else:
            h_deg = 60 * ((r - g) / d) + 240

        pfx = 'oscuro ' if brightness < 0.38 else ('brillante ' if brightness > 0.72 else '')
        if h_deg < 30 or h_deg >= 330:
            return pfx + 'rojo'
        if h_deg < 60:
            return pfx + 'naranja'
        if h_deg < 90:
            return pfx + 'amarillo'
        if h_deg < 150:
            return pfx + 'verde'
        if h_deg < 195:
            return pfx + 'cian'
        if h_deg < 255:
            return pfx + 'azul'
        if h_deg < 285:
            return pfx + 'violeta'
        return pfx + 'magenta'
    except Exception:
        return hex_str


def _add_slide_text(img_bytes, titulo, subtitulo, numero, total, brand_id=None):
    """
    Overlay title + slide indicator on a carousel image using PIL.
    Returns PNG bytes. Always renders text so it's guaranteed to be in Spanish.
    """
    from PIL import Image, ImageDraw
    if not titulo:
        return img_bytes

    img = Image.open(io.BytesIO(img_bytes)).convert('RGBA')
    w, h = img.size
    overlay = Image.new('RGBA', img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    title_size = max(34, h // 16)
    sub_size   = max(22, h // 26)
    ind_size   = max(18, h // 38)
    f_title = _load_font(title_size, brand_id)
    f_sub   = _load_font(sub_size,   brand_id)
    f_ind   = _load_font(ind_size,   brand_id)

    def _wrap(text, max_px, font, max_lines=2):
        words = text.split()
        lines, cur = [], ''
        for word in words:
            test = (cur + ' ' + word).strip()
            try:
                tw = draw.textbbox((0, 0), test, font=font)[2]
            except Exception:
                tw = len(test) * title_size // 2
            if tw <= max_px:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = word
        if cur:
            lines.append(cur)
        return lines[:max_lines]

    pad_x   = int(w * 0.06)
    max_txt = w - pad_x * 2

    title_lines = _wrap(titulo, max_txt, f_title, 2)
    sub_lines   = _wrap(subtitulo, max_txt, f_sub, 2) if subtitulo else []

    lh_title = title_size + 10
    lh_sub   = sub_size + 8
    inner_h  = len(title_lines) * lh_title + (len(sub_lines) * lh_sub + 8 if sub_lines else 0)
    strip_h  = inner_h + 48

    # Gradient-like dark strip at the bottom
    for _yi in range(strip_h + 20):
        alpha = int(220 * min(1.0, (_yi / 20) ** 0.5)) if _yi < 20 else 220
        draw.rectangle([(0, h - strip_h - 20 + _yi), (w, h - strip_h - 19 + _yi)],
                       fill=(10, 12, 24, alpha))

    # Title text
    y = h - strip_h + 20
    for line in title_lines:
        try:
            tw = draw.textbbox((0, 0), line, font=f_title)[2]
        except Exception:
            tw = len(line) * title_size // 2
        x = max(pad_x, (w - tw) // 2)
        draw.text((x + 2, y + 2), line, font=f_title, fill=(0, 0, 0, 180))
        draw.text((x, y), line, font=f_title, fill=(255, 255, 255, 255))
        y += lh_title

    # Subtitle
    if sub_lines:
        y += 4
        for line in sub_lines:
            try:
                tw = draw.textbbox((0, 0), line, font=f_sub)[2]
            except Exception:
                tw = len(line) * sub_size // 2
            x = max(pad_x, (w - tw) // 2)
            draw.text((x, y), line, font=f_sub, fill=(200, 215, 240, 230))
            y += lh_sub

    # Slide indicator badge — top right
    ind = f'{numero} / {total}'
    try:
        ib = draw.textbbox((0, 0), ind, font=f_ind)
        iw, ih = ib[2] - ib[0], ib[3] - ib[1]
    except Exception:
        iw, ih = len(ind) * ind_size // 2, ind_size
    bp = int(w * 0.035)
    bx1, by1 = w - iw - bp * 2 - 4, bp
    bx2, by2 = w - bp, bp + ih + 12
    draw.rectangle([bx1, by1, bx2, by2], fill=(10, 12, 24, 180))
    draw.text((bx1 + bp, by1 + 6), ind, font=f_ind, fill=(200, 215, 240, 230))

    result = Image.alpha_composite(img, overlay)
    out = io.BytesIO()
    result.convert('RGB').save(out, format='PNG')
    return out.getvalue()


def _generate_imagen(prompt_text, aspect_ratio='16:9', quality='standard'):
    """Generates an image and returns raw image bytes."""
    import urllib.request
    key = _get_gemini_key()
    _prompt = prompt_text + _SPANISH_TEXT_SUFFIX

    if quality == 'nano_banana':
        # Gemini 3 Pro Image Preview — generateContent with thinking
        url = (f'https://generativelanguage.googleapis.com/v1beta/models/'
               f'nano-banana-pro-preview:generateContent?key={key}')
        # Add aspect hint to the prompt
        aspect_hint = 'wide landscape format' if aspect_ratio == '16:9' else 'square format'
        body = json.dumps({
            'contents': [{'parts': [{'text': f'{_prompt}, {aspect_hint}'}]}],
            'generationConfig': {'responseModalities': ['IMAGE']},
        }).encode()
        req  = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
        resp = urllib.request.urlopen(req, timeout=120)
        data = json.loads(resp.read())
        parts = data.get('candidates', [{}])[0].get('content', {}).get('parts', [])
        for p in parts:
            if 'inlineData' in p:
                return base64.b64decode(p['inlineData']['data'])
        raise RuntimeError('Nano Banana no devolvió imagen. Intenta de nuevo.')

    else:
        # Imagen 4 — predict endpoint
        _model_map = {
            'fast':     'imagen-4.0-fast-generate-001',
            'standard': 'imagen-4.0-generate-001',
            'ultra':    'imagen-4.0-ultra-generate-001',
        }
        model = _model_map.get(quality, 'imagen-4.0-generate-001')
        url   = (f'https://generativelanguage.googleapis.com/v1beta/models/'
                 f'{model}:predict?key={key}')
        body  = json.dumps({
            'instances':  [{'prompt': _prompt}],
            'parameters': {'sampleCount': 1, 'aspectRatio': aspect_ratio},
        }).encode()
        req  = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
        resp = urllib.request.urlopen(req, timeout=90)
        data = json.loads(resp.read())
        preds = data.get('predictions', [])
        if not preds or 'bytesBase64Encoded' not in preds[0]:
            raise RuntimeError('Imagen 4 no devolvió imagen. Intenta de nuevo.')
        return base64.b64decode(preds[0]['bytesBase64Encoded'])


# ── Dates ──────────────────────────────────────────────────────────────────────

def _prev_month(año, mes):
    return (año - 1, 12) if mes == 1 else (año, mes - 1)


def _pub_dates_in_month(año, mes, dias_semana_list):
    target_wd = set()
    for d in dias_semana_list:
        wd = _DIA_MAP.get(d.lower().strip())
        if wd is not None:
            target_wd.add(wd)
    num_days = cal_mod.monthrange(año, mes)[1]
    return [date(año, mes, day) for day in range(1, num_days + 1)
            if date(año, mes, day).weekday() in target_wd]


# ── Special Dates Research ─────────────────────────────────────────────────────

def _research_special_dates(año, mes, brand_label):
    prompt = f"""Eres un estratega de contenido B2G especializado en tecnología de seguridad pública en México. Tu trabajo es identificar fechas de alto valor real para la comunicación de {brand_label}.

CONTEXTO DE LA MARCA: empresa de tecnología para seguridad pública que vende a gobiernos. Productos: C4/C5, videovigilancia con IA, reconocimiento facial, LPR, centros de mando. Clientes: secretarías de seguridad estatales/municipales, policías, protección civil, gobierno federal.

═══ ANÁLISIS DE {MESES_ES[mes].upper()} {año} ═══

Hay DOS tipos de fechas válidas para esta marca. Evalúa ambos:

──────────────────────────────────────────────
EJE A: TÉCNICO-INSTITUCIONAL
Fechas directamente del sector seguridad pública:
• Días oficiales de cuerpos de seguridad (Día del Policía, Guardia Nacional, Bomberos, Protección Civil)
• Fechas institucionales de SSP, CNSP, SESNSP, C4, C5
• Días relacionados con seguridad ciudadana, prevención del delito, ciudades seguras
• Congresos o expos DOCUMENTADOS del sector (solo si ocurren en {MESES_ES[mes]})

Criterio de aprobación EJE A: ¿Un director de C4/C5 lo pondría en la agenda de su institución? Si sí → incluir.

──────────────────────────────────────────────
EJE B: SOCIAL-EMPÁTICO
Fechas que conectan la marca con el impacto humano de la seguridad pública:
• Días relacionados con víctimas de violencia, feminicidio, desaparición forzada, violencia familiar
• Días de derechos humanos en contextos de seguridad (no derechos humanos genérico)
• Días contra la violencia en general — donde la marca puede hablar de su rol en la PREVENCIÓN y RESPUESTA
• Días de memoria o conmemoración de caídos en servicio (policías, bomberos, rescatistas)

Criterio de aprobación EJE B: ¿La marca puede hablar de este tema con genuina empatía Y mostrar cómo su tecnología contribuye a reducir ese problema en concreto, sin forzar la conexión? Si el contenido sonaría natural y humano → incluir.

──────────────────────────────────────────────
DESCARTE AUTOMÁTICO (aplica a ambos ejes):

✗ ABSURDO TECNOLÓGICO: fechas de otros sectores donde la seguridad pública NO es el tema principal.
  Ejemplos concretos de lo que NO incluir:
  - Día del Océano — aunque alguien argumente que "las cámaras detectan derrames"
  - Día del Medio Ambiente, Día de la Tierra, Día del Agua, Día del Árbol
  - Día de la Salud Mental, Día del Médico, Día de la Educación
  PRUEBA: ¿La seguridad pública, el orden público o la protección ciudadana es el TEMA CENTRAL de esta fecha? Si no → descarta. Si sí → puede incluirse.

✗ FESTIVIDADES GENÉRICAS: Día del Padre, Día de las Madres, Halloween, Navidad, Día del Maestro, Día del Niño — a menos que haya un ángulo explícito de seguridad pública (ej: "Regreso a Clases Seguro" podría aplica si hay campaña de seguridad escolar documentada).

✗ TECNOLOGÍA GENÉRICA: Día del Internet, Día del Programador, Día de la IA sin contexto de seguridad pública gubernamental.

✗ SALUD, EDUCACIÓN, MEDIO AMBIENTE: salvo que el vínculo con seguridad pública sea el tema PRINCIPAL de la fecha, no un ángulo derivado.

✗ CONEXIÓN DE MÁS DE 1 PASO: si para llegar a "seguridad pública" necesitas más de una inferencia, descarta.

──────────────────────────────────────────────
RESULTADO:
• Solo incluye fechas que pasaron el filtro con confianza real
• Máximo 5 en total (EJE A + EJE B combinados). Mínimo 0.
• En "conexion": escribe el ángulo de contenido en máximo 12 palabras — cómo la marca hablaría de este día de forma natural y honesta.
• En "eje": indica "tecnico" o "empatico"

Responde ÚNICAMENTE con JSON válido, sin markdown ni texto previo:
[
  {{
    "fecha": "{año}-{mes:02d}-DD",
    "nombre": "Nombre oficial exacto de la fecha o evento",
    "conexion": "Angulo de contenido en max 12 palabras",
    "tipo": "nacional",
    "relevancia": "alta",
    "eje": "tecnico"
  }}
]

Tipos: "nacional" | "internacional" | "sector"
Relevancia: "alta" (conexión directa y obvia) | "media" (conexión genuina pero requiere buen copy)
Eje: "tecnico" | "empatico"
"""
    provider = _ai_provider()
    try:
        if provider == 'gemini':
            raw_text = _call_gemini_raw(prompt)
        else:
            import anthropic
            client = anthropic.Anthropic(api_key=_get_api_key())
            with client.messages.stream(
                model='claude-opus-4-8',
                max_tokens=8000,
                thinking={"type": "adaptive"},
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                msg = stream.get_final_message()
            raw_text = ""
            for block in msg.content:
                if hasattr(block, 'text') and block.text:
                    raw_text += block.text
    except Exception as e:
        raise RuntimeError(f"Error al conectar con la IA ({provider}): {e}") from e

    raw_text = raw_text.strip()
    start, end = raw_text.find('['), raw_text.rfind(']') + 1
    if start == -1 or end <= 0:
        raise ValueError(f"La IA no devolvió JSON válido. Respuesta:\n{raw_text[:500]}")
    try:
        return json.loads(raw_text[start:end])
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON inválido: {e}\nFragmento: {raw_text[start:start+300]}") from e


# ── Calendar Rendering ─────────────────────────────────────────────────────────

def _render_calendar(año, mes, pub_dates, special_dates):
    pub_set     = {d.strftime('%Y-%m-%d') for d in pub_dates}
    special_set = {sd['fecha'] for sd in special_dates}
    day_names   = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

    cells = ''
    for d in day_names:
        cells += (
            f'<div style="color:#5b8db8;font-size:.68rem;font-weight:700;'
            f'text-align:center;padding:3px 0;">{d}</div>'
        )

    for week in cal_mod.monthcalendar(año, mes):
        for day in week:
            if day == 0:
                cells += '<div></div>'
                continue
            d_str  = f"{año}-{mes:02d}-{day:02d}"
            is_pub = d_str in pub_set
            is_sp  = d_str in special_set
            if is_pub and is_sp:
                bg, color, border = 'linear-gradient(135deg,#1e3a8a 50%,#78350f 50%)', '#fff', '2px solid #f59e0b'
            elif is_pub:
                bg, color, border = '#1e3a5f', '#60a5fa', '1px solid #3b82f6'
            elif is_sp:
                bg, color, border = '#78350f', '#fbbf24', '1px solid #f59e0b'
            else:
                bg, color, border = 'transparent', '#64748b', '1px solid transparent'
            weight = '700' if (is_pub or is_sp) else '400'
            cells += (
                f'<div style="background:{bg};color:{color};border:{border};'
                f'border-radius:5px;padding:5px 2px;font-size:.75rem;font-weight:{weight};'
                f'text-align:center;">{day}</div>'
            )

    return f"""
<div style="background:#0d1b2e;border:1px solid #1e3a5f;border-radius:10px;padding:14px 16px;">
  <div style="color:#60a5fa;font-size:.82rem;font-weight:700;text-align:center;margin-bottom:10px;">
    📅 {MESES_ES[mes]} {año}
  </div>
  <div style="display:grid;grid-template-columns:repeat(7,1fr);gap:3px;">{cells}</div>
  <div style="display:flex;gap:14px;margin-top:10px;font-size:.68rem;color:#8892a4;flex-wrap:wrap;">
    <span><span style="display:inline-block;width:9px;height:9px;background:#1e3a5f;
      border:1px solid #3b82f6;border-radius:2px;margin-right:3px;"></span>Publicación</span>
    <span><span style="display:inline-block;width:9px;height:9px;background:#78350f;
      border:1px solid #f59e0b;border-radius:2px;margin-right:3px;"></span>Fecha especial</span>
  </div>
</div>"""


def _render_status_calendar(año, mes, df):
    """Month calendar — highlights days with publications using status color."""
    pub_days = {}
    for _, row in df.iterrows():
        fecha = str(row.get('Fecha', row.get('fecha', ''))).strip()
        if len(fecha) >= 10:
            key    = fecha[:10]
            estado = _from_display(str(row.get('Estado', row.get('estado', 'Borrador'))))
            if estado not in _ESTADOS:
                estado = 'Borrador'
            if key not in pub_days:
                pub_days[key] = estado  # first post's estado drives the color

    day_names = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    header = ''.join(
        f'<div style="color:#5b8db8;font-size:.7rem;font-weight:700;'
        f'text-align:center;padding:4px 0;">{d}</div>'
        for d in day_names
    )

    cells = ''
    for week in cal_mod.monthcalendar(año, mes):
        for day in week:
            if day == 0:
                cells += '<div></div>'
                continue
            d_str  = f"{año}-{mes:02d}-{day:02d}"
            estado = pub_days.get(d_str)
            if estado:
                c = _ESTADO_COLORS[estado]
                cells += (
                    f'<div style="background:{c["bg"]};border:1.5px solid {c["solid"]};'
                    f'border-radius:5px;padding:3px 2px;text-align:center;">'
                    f'<span style="font-size:.72rem;font-weight:700;color:{c["text"]};">{day}</span>'
                    f'</div>'
                )
            else:
                cells += (
                    f'<div style="border:1px solid #1e2d45;border-radius:5px;'
                    f'padding:3px 2px;text-align:center;">'
                    f'<span style="font-size:.72rem;color:#3d4d60;">{day}</span>'
                    f'</div>'
                )

    legend = ''.join(
        f'<span style="display:inline-flex;align-items:center;gap:5px;font-size:.72rem;color:#8892a4;">'
        f'<span style="width:11px;height:11px;border-radius:50%;background:{_ESTADO_COLORS[e]["solid"]};'
        f'display:inline-block;"></span>{_ESTADO_EMOJI[e]} {e}</span>'
        for e in _ESTADOS
    )

    return f"""
<div style="background:#0d1b2e;border:1px solid #1e3a5f;border-radius:12px;padding:18px 14px;">
  <div style="color:#60a5fa;font-size:.9rem;font-weight:700;text-align:center;margin-bottom:14px;">
    {MESES_ES[mes]} {año}
  </div>
  <div style="display:grid;grid-template-columns:repeat(7,1fr);gap:3px;">
    {header}{cells}
  </div>
  <div style="display:flex;gap:14px;margin-top:14px;flex-wrap:wrap;justify-content:center;">
    {legend}
  </div>
</div>"""


# ── Insights Context ───────────────────────────────────────────────────────────

def _get_ai_insights_from_session(marca_key, prev_año, prev_mes):
    found = {}
    for red in ['LinkedIn', 'Facebook', 'Instagram']:
        key = f"ai_analysis_{marca_key}_{red}_{prev_año}_{prev_mes}"
        if key in st.session_state and st.session_state[key]:
            found[red] = st.session_state[key]
    return found


def _get_raw_insights_context(marca_nombre, prev_año, prev_mes):
    try:
        from database import _conn, get_metricas_mensuales, get_kpi_objetivos, get_publicaciones_count, get_kpi_manual
        from config import METRICAS
        from utils import fmt_num, fmt_pct, kpi_status, mes_nombre

        redes, lines = ['LinkedIn', 'Facebook', 'Instagram'], []
        for red in redes:
            reales = get_metricas_mensuales(marca_nombre, red, prev_año, prev_mes)
            if not reales:
                continue
            objetivos    = get_kpi_objetivos(marca_nombre, red, prev_año, prev_mes)
            metricas_red = METRICAS.get(red, [])
            kpi_lines    = []
            for m in metricas_red:
                tipo = m['tipo']
                if tipo == 'contenido':
                    val = float(get_publicaciones_count(marca_nombre, red, prev_año, prev_mes))
                elif tipo in ('manual', 'manual_50pct'):
                    val = float(get_kpi_manual(marca_nombre, red, m['key'], prev_año, prev_mes) or 0)
                else:
                    val = float(reales.get(m['key'], 0) or 0)
                if val:
                    meta = float(objetivos.get(m['key'], 0) or 0)
                    _, _, pct = kpi_status(val, meta)
                    kpi_lines.append(f"    · {m['label']}: {fmt_num(val)}"
                                     + (f" ({fmt_pct(pct)} cumpl.)" if meta else ""))
            if kpi_lines:
                lines.append(f"  {red}:\n" + "\n".join(kpi_lines))

        with _conn() as con:
            df_fmt = pd.read_sql_query(
                """SELECT tipo, AVG(tasa_interaccion) AS avg_tasa,
                          AVG(reacciones) AS avg_reac, COUNT(*) AS total
                   FROM contenido_posts
                   WHERE marca=? AND strftime('%Y-%m',fecha)=?
                   GROUP BY tipo ORDER BY avg_reac DESC LIMIT 5""",
                con, params=[marca_nombre, f"{prev_año}-{prev_mes:02d}"],
            )

        if not lines and df_fmt.empty:
            return None

        ctx = f"KPIs de {mes_nombre(prev_mes)} {prev_año}:\n" + "\n".join(lines)
        if not df_fmt.empty:
            ctx += "\n\nFormatos con mejor engagement:\n"
            for _, r in df_fmt.iterrows():
                tasa = float(r.get('avg_tasa', 0) or 0)
                ctx += (f"  · {r['tipo']}: {fmt_num(r['avg_reac'])} reac. promedio, "
                        f"tasa {fmt_pct(tasa*100 if tasa<=1 else tasa)} ({int(r['total'])} posts)\n")
        return ctx
    except Exception:
        return None


# ── Prompt Builders ────────────────────────────────────────────────────────────

_PERSONA = """Eres un Community Manager y Creador de Contenido Senior especializado en marcas B2G del sector tecnología y seguridad pública en México, con más de 10 años de experiencia gestionando redes sociales para empresas que venden a gobiernos, secretarías de seguridad y tomadores de decisión del sector público.

Tu expertise incluye:
- Storytelling estratégico para audiencias ejecutivas y gubernamentales
- Hooks que detienen el scroll de gobernadores, secretarios y directores de seguridad
- Estructura narrativa que convierte: gancho → problema → solución → prueba social → CTA
- Copywriting diferenciado por red: LinkedIn (formal, autoridad), Facebook/Instagram (conversacional)
- Conocimiento profundo de C4/C5, videovigilancia, IA aplicada, reconocimiento facial, LPR
- Criterios editoriales para priorizar pilares según rendimiento histórico"""


def _build_brand_section(brand):
    rrss     = brand.get('rrss', {})
    pilares  = rrss.get('pilares', [])
    formatos = rrss.get('formatos', [])
    htags    = rrss.get('hashtags', {})
    cta_list = rrss.get('cta_opciones', [])
    tone     = brand.get('tone', {})

    all_tags   = [t for tags in htags.values() for t in tags]
    marca_tags = htags.get('marca', [])
    pilares_txt = '\n'.join(f"  • {p['label']}: {p.get('descripcion','')}" for p in pilares)

    return f"""
=== BRIEF: {brand.get('label','')} ===
Tagline: "{brand.get('tagline','')}"
Audiencia: {brand.get('audience','')}
Tono: {tone.get('style','')}
Evitar: {', '.join(tone.get('avoid',[]))}
Mensajes clave:
{chr(10).join(f"  • {m}" for m in brand.get('key_messages',[]))}

PILARES:
{pilares_txt}

FORMATOS: {', '.join(formatos)}
HASHTAGS (siempre incluir de marca): marca={' '.join(marca_tags)} | pool={' '.join(all_tags[:16])}
CTAs:
{chr(10).join(f"  • {c}" for c in cta_list)}

REGLA DE ESPEJO:
  • LinkedIn: formal, autoridad, 2-3 párrafos con datos o casos concretos.
  • Facebook/Instagram: COPY IDÉNTICO entre sí, tono conversacional, 1-2 párrafos, emojis moderados."""


def _build_main_prompt(brand, año, mes, objetivo, pub_dates,
                       special_dates_selected, ai_insights, raw_ctx):
    prev_año, prev_mes = _prev_month(año, mes)

    dates_base = '\n'.join(
        f"  {d.strftime('%Y-%m-%d')} ({_DIA_LABEL[d.weekday()]}) — publicación regular"
        for d in pub_dates
    )
    dates_extra = ''
    if special_dates_selected:
        dates_extra = '\n\nFECHAS ESPECIALES (publicaciones EXTRA además de las regulares):\n'
        dates_extra += '\n'.join(
            f"  {sd['fecha']} — {sd['nombre']}: {sd.get('conexion','')}"
            for sd in special_dates_selected
        )

    if ai_insights:
        ctx = (f"\n\n=== ANÁLISIS IA DE {MESES_ES[prev_mes].upper()} {prev_año} ===\n"
               "Usa este análisis como base estratégica para las decisiones de contenido.\n")
        for red, texto in ai_insights.items():
            ctx += f"\n── {red} ──\n{texto[:1400]}{'…' if len(texto)>1400 else ''}\n"
    elif raw_ctx:
        ctx = f"\n\n=== DATOS HISTÓRICOS ({MESES_ES[prev_mes]} {prev_año}) ===\n{raw_ctx}"
    else:
        ctx = ""

    num_base  = len(pub_dates)
    num_extra = len(special_dates_selected)
    num_total = num_base + num_extra

    return f"""{_PERSONA}
{_build_brand_section(brand)}
{ctx}
=== PARRILLA: {MESES_ES[mes].upper()} {año} ===
Objetivo mensual: {objetivo}

DÍAS DE PUBLICACIÓN ({num_base} regulares + {num_extra} especiales = {num_total} piezas total):
{dates_base}{dates_extra}

INSTRUCCIONES:
1. Genera UNA pieza por cada fecha listada arriba (regulares + especiales).
2. Las fechas especiales son publicaciones EXTRA para aprovechar la efeméride — no reemplazan las regulares.
3. Si tienes análisis del mes anterior, prioriza pilares y formatos con mejor rendimiento.
4. Cada post necesita un HOOK potente en la primera línea que detenga el scroll.
5. Arte sugerida: descripción concreta y accionable para el diseñador.
6. Distribuye pilares equilibradamente; si hay datos de rendimiento, pondera según resultados.
7. Selecciona 5-7 hashtags del pool + siempre al menos uno de marca.
8. Alterna formatos entre publicaciones.
9. Para fechas especiales, conecta la efeméride con el mensaje de marca de forma natural.

RESPONDE ÚNICAMENTE con JSON válido — sin markdown, sin texto previo:
[
  {{
    "fecha": "YYYY-MM-DD",
    "dia_semana": "Lunes",
    "tipo_dia": "regular",
    "pilar": "Posicionamiento",
    "formato": "Imagen estática / Infografía",
    "tema": "Título conciso (máx 60 chars)",
    "copy_linkedin": "Copy formal completo para LinkedIn",
    "copy_facebook": "Copy conversacional completo para Facebook e Instagram",
    "arte_sugerencia": "Descripción detallada y accionable para el diseñador",
    "hashtags": "#Tag1 #Tag2 #Tag3 #Tag4 #Tag5",
    "cta": "CTA seleccionada",
    "estado": "Borrador"
  }}
]

tipo_dia puede ser "regular" o "especial".
Genera exactamente {num_total} objetos ordenados por fecha.
"""


def _build_adjustment_prompt(brand, parrilla_json, historial, nuevo_pedido):
    historial_txt = ''
    for i, h in enumerate(historial, 1):
        historial_txt += f"\n[Ajuste #{i}]\nUsuario: {h['pedido']}\nCambios realizados: {h['cambios']}\n"

    return f"""{_PERSONA}
{_build_brand_section(brand)}

=== PARRILLA ACTUAL ===
{json.dumps(parrilla_json, ensure_ascii=False, indent=2)}
{historial_txt}
=== NUEVO PEDIDO DE AJUSTE ===
{nuevo_pedido}

Aplica los cambios solicitados a la parrilla. Modifica SOLO lo necesario — mantén el resto igual.
Respeta siempre el brief de marca, la regla de espejo Facebook/Instagram y la estructura de pilares.
Conserva el campo "estado" de cada post tal como está.

Responde con DOS bloques separados por el separador "---CAMBIOS---":
1. Primero: una descripción breve (2-3 líneas) de qué cambiaste y por qué.
2. Después del separador: el JSON completo de la parrilla ajustada (mismo formato que antes).

Ejemplo:
Modifiqué los posts del 7 y 14 de julio para enfocarlos en Educación Técnica ya que mencionas que Posicionamiento ya está cubierto. Ajusté también los hashtags correspondientes.
---CAMBIOS---
[...JSON completo...]
"""


# ── Image Prompt Builder ───────────────────────────────────────────────────────

def _build_image_prompt_request(row, brand, red_formato, style='libre'):
    tema   = str(row.get('Tema', ''))
    pilar  = str(row.get('Pilar', ''))
    arte   = str(row.get('Arte Sugerida', ''))
    copy   = str(row.get('Copy LinkedIn', ''))[:300]
    fmt    = str(row.get('Formato', ''))
    fecha  = str(row.get('Fecha', ''))
    tipo   = str(row.get('Tipo', 'regular'))
    label  = brand.get('label', '')
    tone   = brand.get('tone', {}).get('style', 'profesional y moderno')
    avoid  = ', '.join(brand.get('tone', {}).get('avoid', []))
    colors = brand.get('colors', {})
    primary   = _hex_to_color_name(colors.get('primary', ''))
    secondary = _hex_to_color_name(colors.get('secondary', ''))
    accent    = _hex_to_color_name(colors.get('accent', colors.get('tertiary', '')))

    if 'LinkedIn' in red_formato:
        aspect = '1200×628 px (horizontal, landscape) — aspect ratio 16:9'
    elif 'Instagram' in red_formato or 'Facebook' in red_formato:
        aspect = '1080×1080 px (cuadrado) — aspect ratio 1:1'
    else:
        aspect = '1200×628 px (LinkedIn, 16:9) y 1080×1080 px (Instagram/Facebook, 1:1)'

    color_section = ''
    if primary:
        color_section = f"""
PALETA DE COLOR DE MARCA (aplicar sutilmente en iluminación, pantallas, reflejos, acentos):
- Color primario: {primary}
- Color secundario: {secondary}
- Color acento: {accent}
Integra estos colores de forma natural — no de manera obvia ni forzada."""

    _style_names = {
        'isometric':   'Isométrico (diorama miniatura de ciudad inteligente)',
        'cards':       'Cards corporativo (tecnología hero + entorno real)',
        'infographic': 'Infografía de ciudad inteligente (digital twin + datos operativos)',
    }

    if style in _style_names:
        style_section = f"""
ESTILO VISUAL SELECCIONADO: {_style_names[style]}
IMPORTANTE: El master prompt de estilo (calidad técnica, materiales, iluminación, restricciones) ya está definido.
Tu tarea es generar ÚNICAMENTE la descripción de escena específica de este post:
- Qué elementos, sujetos u objetos concretos aparecen (tecnología, personas, vehículos, infraestructura)
- La acción o narrativa visual que representa el tema
- La composición y plano (primer plano, vista aérea, panorámica, etc.)
- El ambiente o entorno (ciudad, centro de mando, carretera, edificio gubernamental, etc.)
- Cómo se integran los colores de marca en la escena
NO incluyas referencias a estilo de renderizado, calidad técnica, ni listas de "NO X" — eso lo maneja el master prompt."""
        prompt_word_range = "80-140 palabras en inglés"
    else:
        style_section = ""
        prompt_word_range = "150-250 palabras en inglés, muy detallado: sujeto, composición, iluminación, paleta, estilo, ambiente, acabado técnico"

    return f"""Eres un director de arte especializado en contenido B2G para sector tecnología y seguridad pública en México.

PUBLICACIÓN:
- Marca: {label}
- Fecha: {fecha} · Tipo: {tipo}
- Pilar de contenido: {pilar}
- Formato: {fmt}
- Tema: {tema}
- Sugerencia de arte original: {arte}
- Copy de referencia: {copy}…
- Tono de marca: {tone}
- Evitar: {avoid}
- Formato de imagen destino: {aspect}
{color_section}
{style_section}

REGLAS GENERALES:
- Sin texto ni logos en la imagen
- Evitar imágenes de conflicto, violencia o armas
- Preferir tecnología, profesionalismo, modernidad, entornos urbanos de México/LatAm
- Los prompts deben ser en INGLÉS
- Reflejar visualmente la paleta de color de la marca

Responde SOLO con JSON válido, sin texto antes ni después:
{{
  "descripcion_corta": "(15 palabras máximo describiendo la imagen en español)",
  "dalle3": {{
    "prompt": "({prompt_word_range})",
    "notas": "(1 frase de consejo para usar en ChatGPT)"
  }},
  "gemini": {{
    "prompt": "(80-140 palabras en inglés, descriptivo y directo, incluye colores de marca)",
    "notas": "(1 frase de consejo para usar en Gemini)"
  }}
}}"""


def _build_carousel_slides_prompt(row, brand, n_slides, style='cards'):
    tema      = str(row.get('Tema', ''))
    pilar     = str(row.get('Pilar', ''))
    arte      = str(row.get('Arte Sugerida', ''))
    copy      = str(row.get('Copy LinkedIn', ''))[:300]
    label     = brand.get('label', '')
    tone      = brand.get('tone', {}).get('style', 'profesional y moderno')
    colors    = brand.get('colors', {})
    primary   = _hex_to_color_name(colors.get('primary', ''))
    secondary = _hex_to_color_name(colors.get('secondary', ''))

    _style_names = {
        'cards':       'Cards corporativo — fondo blanco/gris, dispositivo hero prominente, paneles de datos flotantes, entorno real fotorrealista',
        'isometric':   'Isométrico — diorama miniatura de ciudad inteligente con plataforma flotante y luces azules',
        'infographic': 'Infografía digital twin — ciudad con glassmorphism, paneles operativos y flujos de datos',
    }
    style_desc = _style_names.get(style, _style_names['cards'])

    mid_count  = n_slides - 2
    mid_instrs = '\n'.join(
        f'  - Slide {i+2}: CONTENIDO — punto clave {i+1}: dato, capacidad, beneficio o caso de uso concreto'
        for i in range(mid_count)
    )

    return f"""Eres un director de arte y estratega de contenido B2G especializado en tecnología y seguridad pública en México.

PUBLICACIÓN BASE:
- Marca: {label}
- Tema: {pilar} — {tema}
- Copy de referencia: {copy}
- Arte sugerida: {arte}
- Tono: {tone}
- Colores de marca: primario {primary}, secundario {secondary}

TAREA: Diseña la estructura narrativa visual de un carrusel de {n_slides} slides.
Estilo visual: {style_desc}

ESTRUCTURA NARRATIVA:
  - Slide 1: PORTADA — hook visual potente, dato o pregunta que detiene el scroll
{mid_instrs}
  - Slide {n_slides}: CTA — llamado a la acción, contacto, próximo paso

Para CADA slide genera:
- titulo_slide: texto corto (máximo 8 palabras en español) — copy principal visible sobre la imagen
- subtitulo: frase de apoyo (máximo 10 palabras en español) que complementa el título; en slide CTA puede ser el CTA directo
- escena: descripción de la escena visual específica (80-120 palabras en INGLÉS).
  Describe QUÉ mostrar: sujetos, elementos tecnológicos, composición, plano, ambiente, acción narrativa.
  Incluye los colores de marca {primary}/{secondary} como acentos de iluminación o elementos.
  NO incluyas keywords de calidad técnica ni listas "NO X" — eso va en el master prompt aparte.

Responde SOLO con JSON válido, sin texto antes ni después:
{{
  "titulo_carrusel": "(título general, máximo 10 palabras en español)",
  "slides": [
    {{"numero": 1, "tipo": "portada",   "titulo_slide": "...", "subtitulo": "...", "escena": "..."}},
    {{"numero": 2, "tipo": "contenido", "titulo_slide": "...", "subtitulo": "...", "escena": "..."}},
    ...
    {{"numero": {n_slides}, "tipo": "cta", "titulo_slide": "...", "subtitulo": "...", "escena": "..."}}
  ]
}}"""


# ── Claude Calls ───────────────────────────────────────────────────────────────

def _call_claude_json(prompt):
    if _ai_provider() == 'gemini':
        return _call_gemini_raw(prompt)
    import anthropic
    client = anthropic.Anthropic(api_key=_get_api_key())
    with client.messages.stream(
        model='claude-sonnet-4-6',
        max_tokens=8000,
        messages=[{"role": "user", "content": prompt}],
    ) as stream:
        msg = stream.get_final_message()
    for block in msg.content:
        if hasattr(block, 'text') and block.text:
            return block.text
    return ''


def _parse_json(text):
    text = text.strip()
    # Strip markdown code fences if present
    if '```' in text:
        parts = text.split('```')
        for p in parts:
            p = p.strip()
            if p.startswith('json'):
                p = p[4:].strip()
            if p.startswith('['):
                text = p
                break
    start = text.find('[')
    end   = text.rfind(']') + 1
    if start == -1 or end == 0:
        raise ValueError(f"No se encontró JSON.\nRespuesta:\n{text[:500]}")
    return json.loads(text[start:end])


def _parse_json_obj(text):
    """Parse a JSON object {...} from raw text (strips markdown fences)."""
    text = text.strip()
    # Strip markdown code fences if present
    if text.startswith('```'):
        text = text.split('```', 2)[1]
        if text.startswith('json'):
            text = text[4:]
        text = text.strip()
    start = text.find('{')
    end   = text.rfind('}') + 1
    if start == -1 or end == 0:
        raise ValueError(f"No se encontró JSON.\nRespuesta:\n{text[:500]}")
    return json.loads(text[start:end])


def _call_adjustment(prompt):
    """Returns (descripcion_cambios, parrilla_json_list)."""
    provider    = _ai_provider()
    placeholder = st.empty()
    full_text   = ""
    lbl         = "Gemini" if provider == 'gemini' else "Claude"
    thinking_ph = st.info(f"⏳ {lbl} está ajustando la parrilla…")

    if provider == 'gemini':
        for chunk in _stream_gemini(prompt):
            if not full_text:
                thinking_ph.empty()
            full_text += chunk
            placeholder.markdown(full_text + "▌")
    else:
        import anthropic
        client = anthropic.Anthropic(api_key=_get_api_key())
        with client.messages.stream(
            model='claude-sonnet-4-6',
            max_tokens=8000,
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            for chunk in stream.text_stream:
                if not full_text:
                    thinking_ph.empty()
                full_text += chunk
                placeholder.markdown(full_text + "▌")

    placeholder.empty()

    if '---CAMBIOS---' in full_text:
        descripcion, _, json_part = full_text.partition('---CAMBIOS---')
    else:
        descripcion = ""
        json_part   = full_text

    posts = _parse_json(json_part)
    return descripcion.strip(), posts


# ── Excel Export ───────────────────────────────────────────────────────────────

def _to_excel(df):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    buf       = io.BytesIO()
    export_df = df.copy()
    # Drop Estado column from Excel — it's a workflow field
    export_df = export_df.drop(columns=['Estado'], errors='ignore')

    if 'Copy Facebook / Instagram' in export_df.columns:
        export_df.insert(
            export_df.columns.get_loc('Copy Facebook / Instagram') + 1,
            'Copy Instagram',
            export_df['Copy Facebook / Instagram'],
        )

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Parrilla')
        ws = writer.sheets['Parrilla']

        hfill = PatternFill('solid', fgColor='1E3A5F')
        hfont = Font(bold=True, color='FFFFFF', size=10)
        wrap  = Alignment(wrap_text=True, vertical='top')
        wide  = {'Copy LinkedIn', 'Copy Facebook / Instagram', 'Copy Instagram', 'Arte Sugerida'}
        med   = {'Hashtags', 'Tema', 'CTA', 'Pilar', 'Formato'}

        for ci, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
            col[0].fill      = hfill
            col[0].font      = hfont
            col[0].alignment = Alignment(horizontal='center', vertical='center')
            cname = str(col[0].value or '')
            ws.column_dimensions[get_column_letter(ci)].width = (
                55 if cname in wide else 30 if cname in med else 16
            )
            for cell in col[1:]:
                cell.alignment = wrap
        for ri in range(2, ws.max_row + 1):
            ws.row_dimensions[ri].height = 90

    buf.seek(0)
    return buf.read()


# ── DataFrame helpers ──────────────────────────────────────────────────────────

_COL_MAP = {
    'fecha':           'Fecha',
    'dia_semana':      'Día',
    'tipo_dia':        'Tipo',
    'pilar':           'Pilar',
    'formato':         'Formato',
    'tema':            'Tema',
    'copy_linkedin':   'Copy LinkedIn',
    'copy_facebook':   'Copy Facebook / Instagram',
    'arte_sugerencia': 'Arte Sugerida',
    'hashtags':        'Hashtags',
    'cta':             'CTA',
    'estado':          'Estado',
}


def _posts_to_df(posts):
    raw = pd.DataFrame(posts)
    df  = raw.rename(columns={k: v for k, v in _COL_MAP.items() if k in raw.columns})
    for col in _COL_MAP.values():
        if col not in df.columns:
            df[col] = _to_display('Borrador') if col == 'Estado' else ''
    if 'Estado' in df.columns:
        df['Estado'] = df['Estado'].fillna('Borrador').apply(
            lambda x: _to_display(_from_display(x) if x in _DISPLAY_TO_ESTADO
                                   else (x if x in _ESTADOS else 'Borrador'))
        )
    return df[[v for v in _COL_MAP.values() if v in df.columns]]


def _df_to_posts(df):
    inv  = {v: k for k, v in _COL_MAP.items()}
    out  = df.copy()
    # Strip emoji prefix from Estado before saving to DB
    if 'Estado' in out.columns:
        out['Estado'] = out['Estado'].apply(_from_display)
    return out.rename(columns={v: k for v, k in inv.items()}).to_dict('records')


def _posts_from_db(db_posts):
    posts = [{
        'fecha':           p.get('fecha', ''),
        'dia_semana':      p.get('dia_semana', ''),
        'tipo_dia':        p.get('tipo_dia', 'regular'),
        'pilar':           p.get('pilar', ''),
        'formato':         p.get('formato', ''),
        'tema':            p.get('tema', ''),
        'copy_linkedin':   p.get('copy_linkedin', ''),
        'copy_facebook':   p.get('copy_facebook', ''),
        'arte_sugerencia': p.get('arte_sugerencia', ''),
        'hashtags':        p.get('hashtags', ''),
        'cta':             p.get('cta', ''),
        'estado':          p.get('estado', 'Borrador'),
    } for p in db_posts]
    return _posts_to_df(posts)


def _github_sync_db():
    """Commit the SQLite DB to GitHub via Git Data API (supports up to 100 MB).

    Uses only stdlib (urllib.request, base64, json).
    Returns True on success, False on any failure (never raises).
    """
    try:
        import urllib.request
        from database import DB_PATH as _db_path

        _load_env()
        token  = os.environ.get('GITHUB_TOKEN', '').strip()
        repo   = os.environ.get('GITHUB_REPO',  '').strip()
        branch = 'main'
        path   = 'data/rrss.db'
        if not token or not repo or not _db_path.exists():
            return False

        api    = f"https://api.github.com/repos/{repo}"
        hdrs   = {'Authorization': f'token {token}',
                  'Accept': 'application/vnd.github+json',
                  'Content-Type': 'application/json'}
        db_b64 = base64.b64encode(_db_path.read_bytes()).decode('ascii')

        def _req(method, url, body=None):
            data = json.dumps(body).encode() if body else None
            req  = urllib.request.Request(url, data=data, headers=hdrs, method=method)
            with urllib.request.urlopen(req, timeout=60) as r:
                return json.loads(r.read())

        # 1. Create blob
        blob_sha = _req('POST', f"{api}/git/blobs",
                        {"content": db_b64, "encoding": "base64"})["sha"]

        # 2. Get HEAD commit SHA
        head_sha = _req('GET', f"{api}/git/refs/heads/{branch}")["object"]["sha"]

        # 3. Get current tree SHA
        tree_sha = _req('GET', f"{api}/git/commits/{head_sha}")["tree"]["sha"]

        # 4. Create new tree
        new_tree = _req('POST', f"{api}/git/trees",
                        {"base_tree": tree_sha,
                         "tree": [{"path": path, "mode": "100644",
                                   "type": "blob", "sha": blob_sha}]})["sha"]

        # 5. Create commit
        new_commit = _req('POST', f"{api}/git/commits",
                          {"message": "sync: auto-backup RRSS database",
                           "tree": new_tree, "parents": [head_sha],
                           "committer": {"name": "RRSS App", "email": "rrss@kabat.com"}})["sha"]

        # 6. Update branch ref
        _req('PATCH', f"{api}/git/refs/heads/{branch}", {"sha": new_commit})
        return True
    except Exception:
        return False


def _save_df_to_db(df, marca, año, mes):
    from database import save_parrilla_posts
    save_parrilla_posts(marca, año, mes, _df_to_posts(df))
    try:
        _github_sync_db()
    except Exception:
        pass


# ── Monday.com Integration ─────────────────────────────────────────────────────

def _monday_api_key():
    _load_env()
    return os.environ.get('MONDAY_API_KEY', '').strip()


def _monday_board_id():
    _load_env()
    return os.environ.get('MONDAY_BOARD_ID', '').strip()


def _monday_configured():
    return bool(_monday_api_key() and _monday_board_id())


def _monday_request(api_key, query, variables=None):
    import requests as rq
    payload = {"query": query}
    if variables:
        payload["variables"] = variables
    resp = rq.post(
        "https://api.monday.com/v2",
        json=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": api_key,
            "API-Version": "2024-01",
        },
        timeout=30,
    )
    resp.raise_for_status()
    data = resp.json()
    if 'errors' in data:
        raise ValueError(f"Monday.com API error: {data['errors'][0].get('message','?')}")
    return data.get('data', {})


def _monday_get_board_info(api_key, board_id):
    """Returns (board_name, columns_list, groups_list). Raises on error."""
    query = f"""query {{
      boards(ids: [{board_id}]) {{
        name
        columns {{ id title type }}
        groups {{ id title }}
      }}
    }}"""
    data   = _monday_request(api_key, query)
    boards = data.get('boards', [])
    if not boards:
        raise ValueError(f"No se encontró el tablero con ID {board_id}.")
    board  = boards[0]
    return board.get('name', ''), board.get('columns', []), board.get('groups', [])


def _monday_find_group(groups, marca):
    """Returns group_id matching the brand name (brand-only, legacy). or None."""
    marca_lower = marca.lower()
    for g in groups:
        title = g.get('title', '').lower()
        if 'kabat' in marca_lower and 'kabat' in title:
            return g['id']
        if 'sym' in marca_lower and 'sym' in title:
            return g['id']
    return None


def _monday_find_group_month(groups, marca, año, mes):
    """Returns group_id matching brand + month + year, or None."""
    marca_lower = marca.lower()
    mes_lower   = MESES_ES[mes].lower()
    año_str     = str(año)
    for g in groups:
        title = g.get('title', '').lower()
        brand_match = (('kabat' in marca_lower and 'kabat' in title) or
                       ('sym'   in marca_lower and 'sym'   in title))
        month_match = mes_lower in title and año_str in title
        if brand_match and month_match:
            return g['id']
    return None


def _monday_create_group(api_key, board_id, group_name):
    """Creates a new group in the board. Returns group_id or None."""
    mutation = """mutation ($boardId: ID!, $name: String!) {
      create_group(board_id: $boardId, group_name: $name) {
        id title
      }
    }"""
    try:
        data = _monday_request(api_key, mutation, {"boardId": str(board_id), "name": group_name})
        return data.get('create_group', {}).get('id')
    except Exception:
        return None


def _monday_ensure_group(api_key, board_id, groups, marca, año, mes):
    """Find or create the month-specific group. Returns group_id or None."""
    group_id = _monday_find_group_month(groups, marca, año, mes)
    if group_id:
        return group_id
    group_name = f"{marca} · {MESES_ES[mes]} {año}"
    return _monday_create_group(api_key, board_id, group_name)


def _monday_match_columns(columns):
    """Returns dict mapping our field names to {id, type} for each detected column."""
    mapping = {}
    for col in columns:
        title = col['title'].lower()
        ctype = col['type']
        cid   = col['id']
        info  = {'id': cid, 'type': ctype}
        if ctype == 'date' or any(k in title for k in ('fecha', 'date', 'publicación', 'publicacion')):
            mapping.setdefault('fecha', info)
        elif ctype == 'color' or any(k in title for k in ('estado', 'status', 'etapa')):
            mapping.setdefault('estado', info)
        elif 'linkedin' in title:
            mapping.setdefault('copy_linkedin', info)
        elif any(k in title for k in ('facebook', 'fb / instagram', 'instagram')):
            mapping.setdefault('copy_facebook', info)
        elif 'arte' in title:
            mapping.setdefault('arte_sugerencia', info)
        elif 'hashtag' in title:
            mapping.setdefault('hashtags', info)
        elif title == 'cta':
            mapping.setdefault('cta', info)
        elif 'pilar' in title:
            mapping.setdefault('pilar', info)
        elif 'formato' in title:
            mapping.setdefault('formato', info)
    return mapping


def _monday_col_value(field, value, col_type):
    """Formats a value for a Monday.com column based on its type."""
    if not value:
        return None
    if col_type == 'date':
        return {"date": str(value)[:10], "time": "00:00:00"}
    if col_type == 'long_text':
        return {"text": str(value)}
    if col_type == 'color':
        return {"label": str(value)}
    # text and everything else
    return str(value)


def _monday_create_item(api_key, board_id, item_name, col_vals_json, group_id=None):
    if group_id:
        mutation = """mutation ($boardId: ID!, $groupId: String!, $name: String!, $cols: JSON!) {
          create_item(board_id: $boardId, group_id: $groupId, item_name: $name, column_values: $cols) {
            id name
          }
        }"""
        variables = {
            "boardId":  str(board_id),
            "groupId":  str(group_id),
            "name":     item_name,
            "cols":     col_vals_json,
        }
    else:
        mutation = """mutation ($boardId: ID!, $name: String!, $cols: JSON!) {
          create_item(board_id: $boardId, item_name: $name, column_values: $cols) {
            id name
          }
        }"""
        variables = {
            "boardId": str(board_id),
            "name":    item_name,
            "cols":    col_vals_json,
        }
    data = _monday_request(api_key, mutation, variables)
    return data.get('create_item', {}).get('id')


def _monday_add_update(api_key, item_id, body):
    mutation = """mutation ($itemId: ID!, $body: String!) {
      create_update(item_id: $itemId, body: $body) { id }
    }"""
    _monday_request(api_key, mutation, {"itemId": str(item_id), "body": body})


def _monday_push_estados(df, marca, año, mes):
    """After a DB save, push Estado changes to Monday for items that already exist there.
    Returns (updated_count, skipped_count, error_msg_or_None)."""
    if not _monday_configured():
        return 0, 0, None

    from database import get_parrilla_posts
    api_key  = _monday_api_key()
    board_id = _monday_board_id()

    # Get monday_item_ids from DB
    db_posts = get_parrilla_posts(marca, año, mes)
    id_map = {
        (p['fecha'], p.get('tipo_dia', 'regular')): p.get('monday_item_id')
        for p in db_posts
        if p.get('monday_item_id')
    }
    if not id_map:
        return 0, 0, None  # Nothing synced to Monday yet

    # Find the Estado column id on the board
    try:
        _, columns, _ = _monday_get_board_info(api_key, board_id)
    except Exception as e:
        return 0, 0, str(e)

    # Find Estado column: prefer exact title match on text columns first,
    # fall back to any column whose title contains 'estado'
    estado_col = None
    for col in columns:
        if col['title'].lower() == 'estado' and col['type'] == 'text':
            estado_col = {'id': col['id'], 'type': col['type']}
            break
    if not estado_col:
        for col in columns:
            if col['title'].lower() == 'estado':
                estado_col = {'id': col['id'], 'type': col['type']}
                break
    if not estado_col:
        col_map = _monday_match_columns(columns)
        estado_col = col_map.get('estado')
    if not estado_col:
        return 0, 0, (
            "No se encontró columna 'Estado' en Monday. "
            f"Columnas disponibles: {', '.join(c['title'] for c in columns)}"
        )

    col_id   = estado_col['id']
    col_type = estado_col['type']

    if col_type in ('color', 'status'):
        return 0, 0, (
            "La columna 'Estado' en Monday es de tipo Status (con colores nativos). "
            "Solo acepta etiquetas pre-configuradas, por eso 'En diseño' y 'Aprobado' no se sincronizan. "
            "Solución: en Monday, borra esa columna y crea una nueva columna de tipo Texto con el nombre 'Estado'."
        )

    mutation = """mutation ($boardId: ID!, $itemId: ID!, $colId: String!, $val: JSON!) {
      change_column_value(board_id: $boardId, item_id: $itemId, column_id: $colId, value: $val) {
        id
      }
    }"""

    posts      = _df_to_posts(df)
    updated    = 0
    skipped    = 0
    stale_ids  = []

    for p in posts:
        key     = (p.get('fecha', ''), p.get('tipo_dia', 'regular'))
        item_id = id_map.get(key)
        if not item_id:
            skipped += 1
            continue
        estado = p.get('estado', 'Borrador')
        # Format value for the column type
        if col_type == 'long_text':
            val_json = json.dumps({"text": estado})
        elif col_type in ('color', 'status'):
            val_json = json.dumps({"label": estado})
        elif col_type == 'date':
            skipped += 1
            continue
        else:
            # text column: plain JSON-encoded string
            val_json = json.dumps(estado)

        try:
            _monday_request(api_key, mutation, {
                "boardId": str(board_id),
                "itemId":  str(item_id),
                "colId":   col_id,
                "val":     val_json,
            })
            updated += 1
        except Exception as e:
            err_msg = str(e).lower()
            if 'inactive' in err_msg or 'not found' in err_msg or 'invalid' in err_msg:
                # Item was deleted from Monday — clear the stale ID from DB
                stale_ids.append((None, marca, año, mes,
                                  p.get('fecha', ''), p.get('tipo_dia', 'regular')))
            skipped += 1

    if stale_ids:
        from database import update_monday_item_ids
        update_monday_item_ids(stale_ids)

    return updated, skipped, (
        f"{len(stale_ids)} ítem(s) eliminados de Monday — usa '🔄 Re-sincronizar tablero' para crearlos de nuevo."
        if stale_ids else None
    )


def _monday_format_update(p):
    return (
        f"**📅 {p.get('fecha','')} ({p.get('dia_semana','')}) — {p.get('tipo_dia','').upper()}**\n"
        f"**Pilar:** {p.get('pilar','')} | **Formato:** {p.get('formato','')}\n\n"
        f"**🔵 LINKEDIN:**\n{p.get('copy_linkedin','')}\n\n"
        f"**🟣 FACEBOOK / INSTAGRAM:**\n{p.get('copy_facebook','')}\n\n"
        f"**🎨 Arte sugerida:** {p.get('arte_sugerencia','')}\n\n"
        f"**#️⃣ Hashtags:** {p.get('hashtags','')}\n\n"
        f"**📣 CTA:** {p.get('cta','')}"
    )


def _monday_fetch_parrilla(api_key, board_id, marca, año, mes):
    """Read parrilla posts from Monday.com for the given brand/month.
    Returns list of post dicts compatible with _posts_to_df()."""
    board_name, columns, groups = _monday_get_board_info(api_key, board_id)
    col_map  = _monday_match_columns(columns)
    # Prefer month-specific group; fall back to brand-only legacy group
    group_id = (_monday_find_group_month(groups, marca, año, mes) or
                _monday_find_group(groups, marca))

    if not group_id:
        raise ValueError(f"No se encontró grupo para '{marca}' en el tablero de Monday.")

    col_ids_str = ', '.join(f'"{info["id"]}"' for info in col_map.values())

    query = f"""query {{
      boards(ids: [{board_id}]) {{
        groups(ids: ["{group_id}"]) {{
          items_page(limit: 200) {{
            items {{
              id
              name
              column_values(ids: [{col_ids_str}]) {{
                id
                text
                value
              }}
            }}
          }}
        }}
      }}
    }}"""

    data  = _monday_request(api_key, query)
    items = (data
             .get('boards',  [{}])[0]
             .get('groups',  [{}])[0]
             .get('items_page', {})
             .get('items',  []))

    reverse_map = {info['id']: field for field, info in col_map.items()}
    mes_prefix  = f"{año}-{mes:02d}"
    DIAS_ES     = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']

    posts = []
    for item in items:
        cv_by_id = {cv['id']: cv for cv in item.get('column_values', [])}
        vals = {}

        for col_id, cv in cv_by_id.items():
            field    = reverse_map.get(col_id)
            if not field:
                continue
            col_type = col_map[field]['type']
            raw_val  = cv.get('value') or ''
            text_val = cv.get('text') or ''

            if col_type == 'date' and raw_val:
                try:
                    date_obj = json.loads(raw_val)
                    vals[field] = date_obj.get('date', text_val)
                except Exception:
                    vals[field] = text_val[:10] if text_val else ''
            else:
                vals[field] = text_val

        fecha = vals.get('fecha', '')
        if not fecha or not fecha.startswith(mes_prefix):
            continue

        item_name = item.get('name', '')
        tema = item_name.split(' — ', 1)[1].strip() if ' — ' in item_name else item_name

        try:
            from datetime import date as _dparse
            dia_semana = DIAS_ES[_dparse.fromisoformat(fecha).weekday()]
        except Exception:
            dia_semana = ''

        posts.append({
            'fecha':           fecha,
            'dia_semana':      dia_semana,
            'tipo_dia':        'regular',
            'tema':            tema,
            'pilar':           vals.get('pilar', ''),
            'formato':         vals.get('formato', ''),
            'copy_linkedin':   vals.get('copy_linkedin', ''),
            'copy_facebook':   vals.get('copy_facebook', ''),
            'arte_sugerencia': vals.get('arte_sugerencia', ''),
            'hashtags':        vals.get('hashtags', ''),
            'cta':             vals.get('cta', ''),
            'estado':          vals.get('estado', 'Borrador'),
            'monday_item_id':  item.get('id'),
        })

    posts.sort(key=lambda p: p['fecha'])
    return posts


def _monday_sync_parrilla(df, marca, año, mes, api_key, board_id):
    """Creates Monday.com items for each post in df. Returns result dict."""
    from database import get_parrilla_posts, update_monday_item_ids

    # Load existing Monday IDs from DB
    db_posts = get_parrilla_posts(marca, año, mes)
    existing_ids = {
        (p['fecha'], p.get('tipo_dia', 'regular')): p.get('monday_item_id')
        for p in db_posts
    }

    board_name, columns, groups = _monday_get_board_info(api_key, board_id)
    col_map  = _monday_match_columns(columns)
    group_id = _monday_ensure_group(api_key, board_id, groups, marca, año, mes)

    if not group_id:
        raise ValueError(f"No se pudo encontrar o crear el grupo para '{marca}' en Monday.")

    # Field-to-post-key mapping
    _FIELD_MAP = {
        'fecha':           'fecha',
        'estado':          'estado',
        'pilar':           'pilar',
        'formato':         'formato',
        'copy_linkedin':   'copy_linkedin',
        'copy_facebook':   'copy_facebook',
        'arte_sugerencia': 'arte_sugerencia',
        'hashtags':        'hashtags',
        'cta':             'cta',
    }

    posts      = _df_to_posts(df)
    results    = []
    new_ids    = []
    skipped    = 0

    for p in posts:
        fecha    = p.get('fecha', '')
        tipo_dia = p.get('tipo_dia', 'regular')
        key      = (fecha, tipo_dia)

        if existing_ids.get(key):
            skipped += 1
            results.append({'fecha': fecha, 'tipo_dia': tipo_dia,
                            'item_id': existing_ids[key], 'new': False})
            continue

        tema      = p.get('tema', '')
        date_part = fecha[-5:].replace('-', '/') if fecha else ''
        item_name = f"{date_part} — {tema[:50]}"

        # Build column values from all matched columns
        col_vals = {}
        for field, col_info in col_map.items():
            post_key = _FIELD_MAP.get(field)
            if not post_key:
                continue
            raw_val  = p.get(post_key, '')
            if not raw_val:
                continue
            formatted = _monday_col_value(field, raw_val, col_info['type'])
            if formatted is not None:
                col_vals[col_info['id']] = formatted

        col_vals_json = json.dumps(col_vals, ensure_ascii=False) if col_vals else "{}"
        item_id = _monday_create_item(api_key, board_id, item_name, col_vals_json, group_id)

        if item_id:
            _monday_add_update(api_key, item_id, _monday_format_update(p))
            new_ids.append((str(item_id), marca, año, mes, fecha, tipo_dia))
            results.append({'fecha': fecha, 'tipo_dia': tipo_dia,
                            'item_id': str(item_id), 'new': True})

    if new_ids:
        update_monday_item_ids(new_ids)

    return {
        'board_name': board_name,
        'created':    len(new_ids),
        'skipped':    skipped,
        'results':    results,
    }


# ── Main View ──────────────────────────────────────────────────────────────────

def show_parrilla():
    st.title("📅 Parrilla de Contenido")
    st.caption(
        "2 días/semana por red · LinkedIn (formal) · Facebook = Instagram (espejo) · "
        "Fechas especiales como publicaciones extra."
    )

    _user_role = st.session_state.get('current_user', {}).get('role', 'viewer')
    _is_visita = _user_role == 'visita'

    marca_key = st.session_state.get('marca_activa', 'k1')
    marca_id  = 'kabat-one' if marca_key == 'k1' else 'sym-servicios'
    brand     = _load_brand(marca_id)
    if not brand:
        st.error("No se encontró el brief de la marca.")
        return

    # Si la parrilla en sesión es de otra marca, limpiarla para cargar la correcta
    _meta_marca = st.session_state.get('parrilla_meta', {}).get('marca', '')
    if _meta_marca and _meta_marca != brand.get('label', ''):
        for _k in ('parrilla_df', 'parrilla_meta', 'parrilla_historial'):
            st.session_state.pop(_k, None)

    rrss     = brand.get('rrss', {})
    dias_pub = rrss.get('frecuencia', {}).get('dias_publicacion', [])
    pilares  = rrss.get('pilares', [])
    formatos = rrss.get('formatos', [])

    # ── Selector de mes/año ────────────────────────────────────────────────────
    st.markdown("### Configuración")
    hoy = date.today()

    col1, col2 = st.columns([1, 1])
    with col1:
        años = list(range(hoy.year + 2, 2023, -1))
        año  = st.selectbox("Año", años,
                            index=años.index(hoy.year) if hoy.year in años else 0,
                            key="parrilla_año")
    with col2:
        mes = st.selectbox("Mes", list(range(1, 13)),
                           format_func=lambda m: MESES_ES[m],
                           index=hoy.month - 1, key="parrilla_mes")

    pub_dates          = _pub_dates_in_month(año, mes, dias_pub)
    prev_año, prev_mes = _prev_month(año, mes)

    if not pub_dates:
        st.warning(f"No hay días de publicación en {MESES_ES[mes]} {año}. "
                   f"Días configurados: {', '.join(dias_pub)}")
        return

    if _is_visita:
        # Para visita: cargar parrilla directamente desde DB sin mostrar el generador
        try:
            from database import get_parrilla_posts as _gpp_v
            _posts_v = _gpp_v(marca_nombre, año, mes)
            if _posts_v:
                st.session_state['parrilla_df']  = _posts_to_df(_posts_v)
                st.session_state['parrilla_meta'] = {
                    'marca': marca_nombre, 'año': año, 'mes': mes,
                    'objetivo': '', 'brand': brand,
                }
            else:
                st.info(f"No hay parrilla guardada para {MESES_ES[mes]} {año}.")
                return
        except Exception as _ev:
            st.error(f"Error al cargar parrilla: {_ev}")
            return

    # ── Carga desde DB; si está vacía y Monday está configurado, auto-recupera ──
    # Streamlit Community Cloud tiene filesystem efímero — cada reboot borra la DB.
    # Monday.com es el almacenamiento persistente real; la DB es sólo caché de sesión.
    if not _is_visita and 'parrilla_df' not in st.session_state:
        _early_marca = brand.get('label', '')
        try:
            from database import get_parrilla_posts as _gpp_early
            _early_posts = _gpp_early(_early_marca, año, mes)
            if _early_posts:
                st.session_state['parrilla_df']        = _posts_from_db(_early_posts)
                st.session_state['parrilla_historial'] = []
                st.session_state['parrilla_meta'] = {
                    'marca': _early_marca, 'año': año, 'mes': mes,
                    'objetivo': '(guardada en base de datos)',
                    'con_insights': False, 'insights_mes': '', 'brand': brand,
                }
        except Exception:
            pass

    # Auto-recover from Monday silently when DB cache is empty
    if not _is_visita and 'parrilla_df' not in st.session_state and _monday_configured():
        _auto_marca = brand.get('label', '')
        _auto_key   = f"_monday_autoload_{marca_key}_{año}_{mes}"
        if not st.session_state.get(_auto_key):
            with st.spinner(f"Sincronizando parrilla de {MESES_ES[mes]} {año} desde Monday.com…"):
                try:
                    _auto_posts = _monday_fetch_parrilla(
                        _monday_api_key(), _monday_board_id(), _auto_marca, año, mes,
                    )
                    if _auto_posts:
                        _auto_df = _posts_to_df(_auto_posts)
                        _save_df_to_db(_auto_df, _auto_marca, año, mes)
                        st.session_state['parrilla_df']        = _auto_df
                        st.session_state['parrilla_historial'] = []
                        st.session_state['parrilla_meta'] = {
                            'marca': _auto_marca, 'año': año, 'mes': mes,
                            'objetivo': '(sincronizada desde Monday.com)',
                            'con_insights': False, 'insights_mes': '', 'brand': brand,
                        }
                        st.session_state[_auto_key] = True
                        st.rerun()
                    else:
                        st.session_state[_auto_key] = True  # mark as tried so we don't loop
                except Exception:
                    st.session_state[_auto_key] = True

    if not _is_visita and 'parrilla_df' not in st.session_state:
        st.markdown("---")
        st.info(
            f"📭 No se encontró parrilla para **{brand.get('label','')} · {MESES_ES[mes]} {año}** "
            "en la base de datos local.  \n"
            "Si ya sincronizaste esta parrilla con Monday.com puedes recuperarla aquí:"
        )
        _rcol, _ = st.columns([2, 3])
        with _rcol:
            if st.button("📥 Recuperar desde Monday.com",
                         key="btn_monday_recover_top",
                         type="primary", use_container_width=True):
                if not _monday_configured():
                    st.error("Monday.com no está configurado. Agrega MONDAY_API_KEY y MONDAY_BOARD_ID en los secrets.")
                else:
                    # Clear the auto-load flag so it retries
                    st.session_state.pop(f"_monday_autoload_{marca_key}_{año}_{mes}", None)
                    with st.spinner("Importando desde Monday.com…"):
                        try:
                            _rp = _monday_fetch_parrilla(
                                _monday_api_key(), _monday_board_id(),
                                brand.get('label', ''), año, mes,
                            )
                            if not _rp:
                                st.warning(
                                    "No se encontraron ítems en Monday para ese mes y marca. "
                                    "Verifica que el grupo del tablero tenga el nombre de la marca."
                                )
                            else:
                                _rdf = _posts_to_df(_rp)
                                _save_df_to_db(_rdf, brand.get('label', ''), año, mes)
                                st.session_state['parrilla_df']        = _rdf
                                st.session_state['parrilla_historial'] = []
                                st.session_state['parrilla_meta'] = {
                                    'marca': brand.get('label', ''), 'año': año, 'mes': mes,
                                    'objetivo': '(recuperada desde Monday.com)',
                                    'con_insights': False, 'insights_mes': '', 'brand': brand,
                                }
                                st.success(f"✅ {len(_rp)} publicaciones recuperadas desde Monday.com.")
                                st.rerun()
                        except Exception as _re:
                            st.error(f"Error al recuperar: {_re}")
        st.markdown("---")

    # ── Fechas especiales — investigación automática ───────────────────────────
    if not _is_visita:
        st.markdown("---")
        st.markdown("### 📆 Fechas Especiales del Mes")

    fd_key     = f"fechas_esp_{marca_key}_{año}_{mes}"
    fd_sel_key = f"fechas_sel_{marca_key}_{año}_{mes}"

    fechas_ya_buscadas = fd_key in st.session_state

    if not _is_visita:
        col_inv, col_ref, col_info = st.columns([1.8, 1, 3])
        with col_inv:
            lbl_buscar = "✅ Fechas buscadas" if fechas_ya_buscadas else "🔍  Buscar fechas del mes"
            if st.button(lbl_buscar, use_container_width=True,
                         key="btn_investigar_fechas",
                         disabled=fechas_ya_buscadas,
                         type="secondary" if fechas_ya_buscadas else "primary"):
                with st.spinner(f"Analizando {MESES_ES[mes]} {año}…"):
                    try:
                        fechas = _research_special_dates(año, mes, brand.get('label', ''))
                        st.session_state[fd_key]     = fechas
                        st.session_state[fd_sel_key] = [f['fecha'] for f in fechas]
                    except Exception as e:
                        st.error(f"Error al buscar fechas: {e}")
                        st.stop()
                st.rerun()
        with col_ref:
            if fechas_ya_buscadas:
                if st.button("🔄 Regenerar", use_container_width=True,
                             key="btn_regenerar_fechas",
                             help="Vuelve a buscar fechas (puede dar resultados distintos)"):
                    st.session_state.pop(fd_key, None)
                    st.session_state.pop(fd_sel_key, None)
                    st.rerun()
        with col_info:
            if fechas_ya_buscadas:
                n = len(st.session_state.get(fd_key, []))
                st.caption(
                    f"{'✅' if n > 0 else 'ℹ️'} {'Se encontraron' if n > 0 else 'No se encontraron'} "
                    f"**{n} fechas** relevantes para el sector en {MESES_ES[mes]}. "
                    f"Solo se incluyen fechas con vínculo directo a seguridad pública."
                )
            else:
                st.caption(
                    "Análisis específico del sector: solo fechas con conexión directa a "
                    "C4/C5, videovigilancia o tecnología de mando. Calidad sobre cantidad."
                )

        special_dates   = st.session_state.get(fd_key, [])
        selected_fechas = st.session_state.get(fd_sel_key, [])

        if special_dates:
            col_cal, col_list = st.columns([1, 1.4])
            with col_cal:
                st.markdown(
                    _render_calendar(año, mes, pub_dates, special_dates),
                    unsafe_allow_html=True,
                )
            with col_list:
                altas  = [sd for sd in special_dates if sd.get('relevancia','alta') == 'alta']
                medias = [sd for sd in special_dates if sd.get('relevancia','') == 'media']
                new_sel = list(selected_fechas)

                _eje_icon = {"tecnico": "🔒", "empatico": "🤝"}

                if altas:
                    st.markdown("**🎯 Alta relevancia**")
                    for sd in altas:
                        d_obj   = date.fromisoformat(sd['fecha'])
                        dia_lbl = f"{d_obj.day} {MESES_ES[mes][:3]} ({_DIA_LABEL[d_obj.weekday()]})"
                        tipo_b  = {"nacional": "🇲🇽", "internacional": "🌍", "sector": "🔒"}.get(sd.get('tipo',''), "📌")
                        eje_b   = _eje_icon.get(sd.get('eje', 'tecnico'), '🔒')
                        checked = sd['fecha'] in selected_fechas
                        val = st.checkbox(
                            f"{tipo_b}{eje_b} **{dia_lbl}** — {sd['nombre']}",
                            value=checked,
                            help=f"💡 {sd.get('conexion', '')}",
                            key=f"fd_chk_{sd['fecha']}",
                        )
                        if val and sd['fecha'] not in new_sel:
                            new_sel.append(sd['fecha'])
                        elif not val and sd['fecha'] in new_sel:
                            new_sel.remove(sd['fecha'])

                if medias:
                    st.markdown("**📌 Relevancia media**")
                    for sd in medias:
                        d_obj   = date.fromisoformat(sd['fecha'])
                        dia_lbl = f"{d_obj.day} {MESES_ES[mes][:3]} ({_DIA_LABEL[d_obj.weekday()]})"
                        tipo_b  = {"nacional": "🇲🇽", "internacional": "🌍", "sector": "🔒"}.get(sd.get('tipo',''), "📌")
                        eje_b   = _eje_icon.get(sd.get('eje', 'tecnico'), '🔒')
                        checked = sd['fecha'] in selected_fechas
                        val = st.checkbox(
                            f"{tipo_b}{eje_b} {dia_lbl} — {sd['nombre']}",
                            value=checked,
                            help=f"💡 {sd.get('conexion', '')}",
                            key=f"fd_chk_{sd['fecha']}",
                        )
                        if val and sd['fecha'] not in new_sel:
                            new_sel.append(sd['fecha'])
                        elif not val and sd['fecha'] in new_sel:
                            new_sel.remove(sd['fecha'])

                if not altas and not medias:
                    st.info(
                        f"No se encontraron fechas con conexión directa al sector "
                        f"en {MESES_ES[mes]}. Puedes usar **🔄 Regenerar** o continuar "
                        f"solo con los días regulares de publicación."
                    )

                st.session_state[fd_sel_key] = new_sel
                selected_fechas = new_sel
        else:
            st.markdown(
                _render_calendar(año, mes, pub_dates, []),
                unsafe_allow_html=True,
            )
            if not fechas_ya_buscadas:
                st.caption(
                    "Pulsa **Buscar fechas del mes** para que Claude identifique "
                    "efemérides específicas del sector seguridad pública."
                )

        special_dates_selected = [sd for sd in special_dates if sd['fecha'] in selected_fechas]

        # ── Objetivo mensual ───────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 🎯 Objetivo del Mes")

        obj_preset = st.selectbox("Selecciona el objetivo principal", OBJETIVOS_PRESET,
                                   key="parrilla_obj_preset")
        objetivo = ""
        if obj_preset == "Personalizado…":
            objetivo = st.text_input("Describe el objetivo mensual",
                                     placeholder="Ej: Posicionar en el sector público del bajío…",
                                     key="parrilla_obj_custom")
        else:
            objetivo = obj_preset
            st.caption(f"Objetivo seleccionado: **{objetivo}**")

        # ── Resumen del mes ────────────────────────────────────────────────────────
        st.markdown("---")
        ai_insights  = _get_ai_insights_from_session(marca_key, prev_año, prev_mes)
        total_piezas = len(pub_dates) + len(special_dates_selected)

        if ai_insights:
            st.success(
                f"**{brand.get('label','')}** · **{MESES_ES[mes]} {año}**  →  "
                f"{len(pub_dates)} publicaciones regulares + {len(special_dates_selected)} especiales "
                f"= **{total_piezas} piezas**  \n"
                f"✅ Insights de IA de **{MESES_ES[prev_mes]} {prev_año}** disponibles "
                f"({', '.join(ai_insights.keys())}) — se usarán como base estratégica."
            )
        else:
            st.info(
                f"**{brand.get('label','')}** · **{MESES_ES[mes]} {año}**  →  "
                f"{len(pub_dates)} publicaciones regulares + {len(special_dates_selected)} especiales "
                f"= **{total_piezas} piezas**  \n"
                f"ℹ️ Sin análisis de Insights guardado para {MESES_ES[prev_mes]} {prev_año}. "
                "Para mejor resultado, genera primero el análisis en **🤖 Insights para IA**."
            )

        # ── Botones principales ────────────────────────────────────────────────────
        col_gen, col_clear = st.columns([3, 1])
        with col_gen:
            generate = st.button(
                "✨  Generar Parrilla con IA",
                type="primary", use_container_width=True,
                disabled=not objetivo.strip(),
                key="btn_generar_parrilla",
            )
        with col_clear:
            if st.button("🗑️  Limpiar", use_container_width=True, key="btn_limpiar_parrilla"):
                for k in ('parrilla_df', 'parrilla_meta', 'parrilla_historial'):
                    st.session_state.pop(k, None)
                st.rerun()

        if not objetivo.strip():
            st.caption("Selecciona o escribe un objetivo para habilitar la generación.")

        # ── Generación ─────────────────────────────────────────────────────────────
        if generate and objetivo.strip():
            with st.status("Generando parrilla con IA…", expanded=True) as status:
                raw_ctx = None
                if ai_insights:
                    st.write(f"Cargando análisis de Insights de {MESES_ES[prev_mes]} {prev_año}…")
                else:
                    st.write(f"Consultando datos de {MESES_ES[prev_mes]} {prev_año}…")
                    raw_ctx = _get_raw_insights_context(brand.get('label', ''), prev_año, prev_mes)

                n_esp = len(special_dates_selected)
                st.write(f"Construyendo {len(pub_dates)} piezas regulares"
                         + (f" + {n_esp} especiales" if n_esp else "") + "…")

                prompt = _build_main_prompt(
                    brand, año, mes, objetivo, pub_dates,
                    special_dates_selected, ai_insights, raw_ctx,
                )

                _lbl = "Gemini 2.5 Flash" if _ai_provider() == 'gemini' else "Claude Sonnet"
                st.write(f"Llamando a {_lbl} (20-60 segundos)…")
                try:
                    raw   = _call_claude_json(prompt)
                    posts = _parse_json(raw)
                    df    = _posts_to_df(posts)

                    st.session_state['parrilla_df']        = df
                    st.session_state['parrilla_historial'] = []
                    st.session_state['parrilla_meta'] = {
                        'marca':        brand.get('label', ''),
                        'año':          año,
                        'mes':          mes,
                        'objetivo':     objetivo,
                        'con_insights': bool(ai_insights),
                        'insights_mes': f"{MESES_ES[prev_mes]} {prev_año}",
                        'brand':        brand,
                    }
                    # Guardar en DB automáticamente
                    try:
                        _save_df_to_db(df, brand.get('label', ''), año, mes)
                        st.write("✅ Guardado en base de datos.")
                    except Exception as _dbe:
                        st.warning(f"⚠️ No se pudo guardar en DB: {_dbe}")

                    status.update(label=f"Parrilla lista: {len(df)} piezas", state="complete")
                    st.rerun()
                except Exception as e:
                    status.update(label="Error al generar", state="error")
                    st.error(f"Error: {e}")

        # ── Intentar cargar desde DB si no hay en sesión ───────────────────────────
        if 'parrilla_df' not in st.session_state:
            marca_nombre = brand.get('label', '')
            try:
                from database import get_parrilla_posts
                db_posts = get_parrilla_posts(marca_nombre, año, mes)
                if db_posts:
                    df_loaded = _posts_from_db(db_posts)
                    st.session_state['parrilla_df']        = df_loaded
                    st.session_state['parrilla_historial'] = []
                    st.session_state['parrilla_meta'] = {
                        'marca': marca_nombre, 'año': año, 'mes': mes,
                        'objetivo': '(guardada en base de datos)',
                        'con_insights': False, 'insights_mes': '', 'brand': brand,
                    }
                    st.info(f"📂 Parrilla de **{MESES_ES[mes]} {año}** cargada desde la base de datos.")
            except Exception:
                pass

        # ── Recuperar desde Monday si la DB está vacía y Monday está configurado ──
        if 'parrilla_df' not in st.session_state and _monday_configured():
            st.markdown("---")
            st.info(
                f"📭 No hay parrilla guardada para **{brand.get('label','')}** "
                f"· **{MESES_ES[mes]} {año}** en la base de datos local.  \n"
                "Si ya sincronizaste esta parrilla con Monday.com, puedes recuperarla aquí."
            )
            if st.button("📥 Recuperar parrilla desde Monday.com",
                         key="btn_monday_recover", type="primary",
                         use_container_width=False):
                with st.spinner("Importando desde Monday.com…"):
                    try:
                        _rec_api     = _monday_api_key()
                        _rec_board   = _monday_board_id()
                        _rec_marca   = brand.get('label', '')
                        _rec_posts   = _monday_fetch_parrilla(_rec_api, _rec_board, _rec_marca, año, mes)
                        if not _rec_posts:
                            st.warning(
                                f"No se encontraron ítems para {_rec_marca} en "
                                f"{MESES_ES[mes]} {año}. Verifica que el grupo en Monday "
                                "tenga el nombre de la marca y que haya ítems para ese mes."
                            )
                        else:
                            _rec_df = _posts_to_df(_rec_posts)
                            _save_df_to_db(_rec_df, _rec_marca, año, mes)
                            st.session_state['parrilla_df']        = _rec_df
                            st.session_state['parrilla_historial'] = []
                            st.session_state['parrilla_meta'] = {
                                'marca': _rec_marca, 'año': año, 'mes': mes,
                                'objetivo': '(recuperada desde Monday.com)',
                                'con_insights': False, 'insights_mes': '', 'brand': brand,
                            }
                            st.success(f"✅ {len(_rec_posts)} publicaciones recuperadas desde Monday.com.")
                            st.rerun()
                    except Exception as _re:
                        st.error(f"Error al recuperar: {_re}")

    if 'parrilla_df' not in st.session_state:
        return

    # ── Header de la parrilla ──────────────────────────────────────────────────
    meta      = st.session_state.get('parrilla_meta', {})
    mes_label = MESES_ES.get(meta.get('mes', 1), '')
    historial = st.session_state.get('parrilla_historial', [])

    st.markdown(f"### {meta.get('marca','')} · {mes_label} {meta.get('año','')}")
    badge = (f"✅ Basada en Insights de **{meta.get('insights_mes','')}**"
             if meta.get('con_insights')
             else f"ℹ️ Basada en datos de **{meta.get('insights_mes','')}**"
             if meta.get('insights_mes')
             else "📂 Cargada desde base de datos")
    st.caption(f"Objetivo: {meta.get('objetivo','')}  ·  {badge}")

    # Normalize Estado column to display format (handles legacy clean values from DB/old sessions)
    _df_norm = st.session_state['parrilla_df']
    if 'Estado' in _df_norm.columns:
        _df_norm = _df_norm.copy()
        _df_norm['Estado'] = _df_norm['Estado'].apply(
            lambda x: x if x in _ESTADOS_DISPLAY
                      else _to_display(_from_display(x) if x in _DISPLAY_TO_ESTADO
                                       else (x if x in _ESTADOS else 'Borrador'))
        )
        st.session_state['parrilla_df'] = _df_norm

    df = st.session_state['parrilla_df']

    # ── TABS ───────────────────────────────────────────────────────────────────
    if _is_visita:
        tab1, tab2 = st.tabs(["📋 Parrilla", "📅 Calendario"])
        tab3 = tab4 = None
    else:
        tab1, tab2, tab3, tab4 = st.tabs(["📋 Parrilla", "📅 Calendario", "🟦 Monday.com", "🎨 Prompts de Imagen"])

    # ═══ TAB 1: PARRILLA ════════════════════════════════════════════════════════
    with tab1:
        pilar_opts   = [p['label'] for p in pilares] if pilares else None
        formato_opts = formatos if formatos else None

        # ── Vista compacta (sin scroll horizontal) ───────────────────────────
        _compact_cols = [c for c in ['Fecha', 'Día', 'Tipo', 'Tema', 'Pilar', 'Formato', 'Estado']
                         if c in df.columns]
        compact_df = df[_compact_cols].copy()

        if not _is_visita:
            compact_df.insert(0, '🗑️', False)

        _ccfg: dict = {
            'Tema':   st.column_config.TextColumn(width='medium'),
            'Tipo':   st.column_config.SelectboxColumn(options=['regular', 'especial'], width='small'),
            'Estado': st.column_config.SelectboxColumn(options=_ESTADOS_DISPLAY, width='small'),
        }
        if not _is_visita:
            _ccfg['🗑️'] = st.column_config.CheckboxColumn("🗑️", width='small',
                                                            help="Marcar para eliminar")
        if pilar_opts:
            _ccfg['Pilar']   = st.column_config.SelectboxColumn(options=pilar_opts,   width='medium')
        if formato_opts:
            _ccfg['Formato'] = st.column_config.SelectboxColumn(options=formato_opts, width='medium')

        _row_h = min(38 * max(len(compact_df), 1) + 42, 500)

        if _is_visita:
            st.dataframe(compact_df, use_container_width=True, height=_row_h)
            compact_edited = compact_df.copy()
        else:
            compact_edited = st.data_editor(
                compact_df, use_container_width=True, num_rows='fixed',
                column_config=_ccfg, height=_row_h, key='parrilla_compact',
            )

        # Sincronizar cambios de la vista compacta de vuelta al df completo
        for _col in _compact_cols:
            if _col in compact_edited.columns and _col in df.columns:
                df[_col] = compact_edited[_col].values

        edited_df = df
        st.session_state['parrilla_df'] = edited_df

        # ── Confirmar y ejecutar eliminación ─────────────────────────────────
        if not _is_visita and '🗑️' in compact_edited.columns:
            _del_mask   = compact_edited['🗑️'] == True
            _del_rows   = compact_edited[_del_mask]
            if not _del_rows.empty:
                _del_fechas = _del_rows['Fecha'].astype(str).tolist()
                _del_tipos  = (_del_rows['Tipo'].astype(str).tolist()
                               if 'Tipo' in _del_rows.columns
                               else ['regular'] * len(_del_rows))
                st.warning(
                    f"⚠️ Vas a eliminar **{len(_del_rows)} post(s)**: "
                    + ', '.join(_del_fechas)
                )
                _cy, _cn, _ = st.columns([1.2, 1, 4])
                if _cy.button("🗑️ Confirmar eliminación", type="primary",
                               key="btn_del_confirm"):
                    from database import delete_parrilla_post as _del_fn
                    _marca_del = meta.get('marca', '')
                    for _fd, _td in zip(_del_fechas, _del_tipos):
                        _del_fn(_marca_del, año, mes, _fd, _td)
                    _github_sync_db()
                    from database import get_parrilla_posts as _gpp_del
                    _new_posts = _gpp_del(_marca_del, año, mes)
                    st.session_state['parrilla_df'] = _posts_to_df(_new_posts)
                    st.success(f"✅ {len(_del_rows)} post(s) eliminado(s).")
                    st.rerun()
                if _cn.button("Cancelar", key="btn_del_cancel"):
                    st.rerun()

        # ── Editor de contenido completo (sin scroll horizontal) ─────────────
        if not _is_visita and len(df) > 0:
            st.markdown("")
            with st.expander("✏️  Ver / editar contenido completo de un post"):
                _pnum = st.number_input(
                    "Número de post en la lista (1 = el primero)",
                    min_value=1, max_value=len(df), value=1, step=1,
                    key="par_detail_sel",
                )
                _pi   = _pnum - 1
                _prow = df.iloc[_pi]

                st.caption(
                    f"Post #{_pnum} — **{_prow.get('Fecha','')}** · {_prow.get('Día','')}"
                )
                _dc1, _dc2 = st.columns(2)
                _d_tema = _dc1.text_input("Tema",
                                          value=str(_prow.get('Tema',    '')),
                                          key=f"d_tema_{_pi}")
                _d_cta  = _dc2.text_input("CTA",
                                          value=str(_prow.get('CTA',     '')),
                                          key=f"d_cta_{_pi}")
                _d_li   = st.text_area("Copy LinkedIn",
                                       value=str(_prow.get('Copy LinkedIn', '')),
                                       height=220, key=f"d_li_{_pi}")
                _d_fb   = st.text_area("Copy Facebook / Instagram",
                                       value=str(_prow.get('Copy Facebook / Instagram', '')),
                                       height=220, key=f"d_fb_{_pi}")
                _d_ht   = st.text_input("Hashtags",
                                        value=str(_prow.get('Hashtags', '')),
                                        key=f"d_ht_{_pi}")
                _d_arte = st.text_area("Arte Sugerida",
                                       value=str(_prow.get('Arte Sugerida', '')),
                                       height=100, key=f"d_arte_{_pi}")

                if st.button("💾  Actualizar este post", type="primary",
                             key=f"d_save_{_pi}", use_container_width=True):
                    df.at[_pi, 'Tema']                      = _d_tema
                    df.at[_pi, 'CTA']                       = _d_cta
                    df.at[_pi, 'Copy LinkedIn']             = _d_li
                    df.at[_pi, 'Copy Facebook / Instagram'] = _d_fb
                    df.at[_pi, 'Hashtags']                  = _d_ht
                    df.at[_pi, 'Arte Sugerida']             = _d_arte
                    edited_df = df
                    st.session_state['parrilla_df'] = df
                    st.success(
                        "Post actualizado en memoria. Presiona **💾 Guardar cambios** "
                        "para persistir en la base de datos."
                    )

        # Botones de guardar y descargar
        col_save, col_dl, col_info = st.columns([1.5, 1.5, 3])
        with col_save:
            if not _is_visita and st.button("💾  Guardar cambios", use_container_width=True, key="btn_guardar_parrilla"):
                try:
                    _save_df_to_db(edited_df, meta.get('marca', ''), año, mes)
                    # Sync DB to GitHub for persistence across Streamlit Cloud reboots
                    _gh_ok = _github_sync_db()
                    if _gh_ok:
                        st.caption("☁️ DB sincronizada con GitHub")
                    # Sync estados to Monday if configured
                    updated, skipped, err = _monday_push_estados(
                        edited_df, meta.get('marca', ''), año, mes
                    )
                    if err:
                        st.warning(f"Guardado en DB ✓ — Monday: {err}")
                    elif updated > 0:
                        st.success(
                            f"Guardado ✓ · {updated} estado(s) actualizados en Monday.com"
                            + (f" · {skipped} sin ítem en Monday" if skipped else "")
                        )
                    elif _monday_configured():
                        from database import get_parrilla_posts as _gpp2
                        _has_ids = any(
                            p.get('monday_item_id')
                            for p in _gpp2(meta.get('marca', ''), año, mes)
                        )
                        if _has_ids:
                            st.warning(
                                "Guardado en DB ✓ — Monday no actualizó (0 ítems). "
                                "Revisa el tab 🟦 → 🔍 Diagnóstico de columnas."
                            )
                        else:
                            st.success(
                                "Guardado en DB ✓ — Monday: sincroniza primero en el tab 🟦"
                            )
                    else:
                        st.success("Cambios guardados en la base de datos.")
                except Exception as e:
                    st.error(f"Error al guardar: {e}")
        with col_dl:
            nombre = (f"Parrilla_{meta.get('marca','').replace(' ','_')}_"
                      f"{mes_label}{meta.get('año','')}.xlsx")
            st.download_button(
                "📥  Descargar Excel",
                data=_to_excel(edited_df),
                file_name=nombre,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_info:
            n = len(edited_df)
            st.caption(
                f"{n} piezas · LinkedIn + Facebook + Instagram · "
                "El Excel incluye columna Instagram separada (espejo de Facebook). "
                "**Cambia el Estado** para dar seguimiento al flujo de trabajo."
            )

        # Semáforo de estados
        estado_counts = {e: 0 for e in _ESTADOS}
        for _, row in edited_df.iterrows():
            e = _from_display(str(row.get('Estado', 'Borrador')))
            if e in estado_counts:
                estado_counts[e] += 1
        total = max(sum(estado_counts.values()), 1)

        cards_html = '<div style="display:flex;gap:10px;margin:12px 0;">'
        for estado in _ESTADOS:
            c   = _ESTADO_COLORS[estado]
            n   = estado_counts[estado]
            pct = int(n / total * 100)
            emoji = _ESTADO_EMOJI[estado]
            cards_html += f"""
            <div style="flex:1;background:{c['bg']};border:1.5px solid {c['solid']};
                        border-radius:10px;padding:14px 12px;text-align:center;">
              <div style="font-size:1.6rem;line-height:1;">{emoji}</div>
              <div style="color:{c['text']};font-weight:700;font-size:1.3rem;
                          margin:4px 0 2px;">{n}</div>
              <div style="color:{c['text']};font-size:.75rem;opacity:.9;">{estado}</div>
              <div style="margin-top:6px;background:#0d1117;border-radius:4px;height:4px;">
                <div style="width:{pct}%;background:{c['solid']};height:4px;
                             border-radius:4px;"></div>
              </div>
            </div>"""
        cards_html += '</div>'
        st.markdown(cards_html, unsafe_allow_html=True)

        # ── Cuadro de ajustes (solo usuarios con permisos de edición) ─────────
        if _is_visita:
            st.info("👤 Modo lectura — contacta al administrador para solicitar cambios.")
        else:
            st.markdown("---")
            st.markdown("### 💬 Ajustar Parrilla")
            st.caption(
                "Dile a Claude qué quieres cambiar: temas ya cubiertos, eventos próximos, "
                "tono, redistribución de pilares, o cualquier otro ajuste."
            )

            if historial:
                with st.expander(f"Historial de ajustes ({len(historial)})", expanded=False):
                    for i, h in enumerate(historial, 1):
                        st.markdown(f"**Ajuste #{i}:** {h['pedido']}")
                        st.markdown(f"*Claude:* {h['cambios']}")
                        st.markdown("---")

            pedido = st.text_area(
                "¿Qué quieres cambiar?",
                placeholder=(
                    "Ej: Los temas de posicionamiento ya los tengo cubiertos con 3 publicaciones. "
                    "Cámbiame esos posts por educación técnica o casos de uso. "
                    "Además, el post del lunes 7 hazlo sobre el lanzamiento del nuevo módulo de LPR."
                ),
                height=110,
                key="parrilla_pedido_ajuste",
            )

            col_ajustar, col_sp = st.columns([2, 3])
            with col_ajustar:
                ajustar = st.button(
                    "🔄  Aplicar Ajuste",
                    type="primary", use_container_width=True,
                    disabled=not pedido.strip(),
                    key="btn_aplicar_ajuste",
                )

            if ajustar and pedido.strip():
                current_posts = _df_to_posts(st.session_state['parrilla_df'])
                adj_prompt    = _build_adjustment_prompt(
                    meta.get('brand', brand), current_posts, historial, pedido
                )
                try:
                    descripcion, new_posts = _call_adjustment(adj_prompt)
                    new_df = _posts_to_df(new_posts)
                    st.session_state['parrilla_df'] = new_df
                    historial.append({'pedido': pedido, 'cambios': descripcion or "Parrilla ajustada."})
                    st.session_state['parrilla_historial'] = historial
                    try:
                        _save_df_to_db(new_df, meta.get('marca', ''), año, mes)
                    except Exception:
                        pass
                    if descripcion:
                        st.success(f"**Cambios realizados:** {descripcion}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al ajustar: {e}")

    # ═══ TAB 2: CALENDARIO ══════════════════════════════════════════════════════
    with tab2:
        st.markdown(
            _render_status_calendar(año, mes, st.session_state['parrilla_df']),
            unsafe_allow_html=True,
        )
        st.caption(
            "El color de cada día refleja el estado actual del post. "
            "Cambia el estado en la pestaña **📋 Parrilla** y pulsa **💾 Guardar cambios**."
        )

        # Listado compacto con estado por pieza
        st.markdown("#### Resumen por publicación")
        df_cal = st.session_state['parrilla_df']
        for _, row in df_cal.iterrows():
            fecha   = str(row.get('Fecha', ''))
            dia     = str(row.get('Día', ''))
            tipo    = str(row.get('Tipo', 'regular'))
            tema    = str(row.get('Tema', ''))
            pilar   = str(row.get('Pilar', ''))
            estado  = str(row.get('Estado', 'Borrador'))
            c       = _ESTADO_COLORS.get(estado, _ESTADO_COLORS['Borrador'])

            label_tipo = "⭐ Especial" if tipo == 'especial' else "📅 Regular"
            emoji      = _ESTADO_EMOJI.get(estado, '⚪')
            badge_html = (
                f'<span style="background:{c["badge"]};color:{c["text"]};border:1px solid {c["solid"]};'
                f'border-radius:6px;padding:3px 10px;font-size:.75rem;font-weight:700;'
                f'white-space:nowrap;">'
                f'{emoji} {estado}</span>'
            )
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:10px;'
                f'padding:6px 0;border-bottom:1px solid #1a2442;">'
                f'<span style="color:#5b8db8;font-size:.78rem;min-width:100px;">{fecha}</span>'
                f'<span style="color:#8892a4;font-size:.75rem;min-width:60px;">{dia}</span>'
                f'<span style="color:#60a5fa;font-size:.75rem;min-width:80px;">{label_tipo}</span>'
                f'<span style="color:#dde3ee;font-size:.78rem;flex:1;">{tema}</span>'
                f'<span style="color:#8892a4;font-size:.72rem;min-width:80px;">{pilar}</span>'
                f'{badge_html}</div>',
                unsafe_allow_html=True,
            )

    # ═══ TAB 3: MONDAY.COM ══════════════════════════════════════════════════════
    if not _is_visita:
     with tab3:
        st.markdown("#### 🟦 Sincronizar con Monday.com")
        st.caption(
            "Exporta cada publicación de la parrilla como un ítem en tu tablero de Monday.com. "
            "El contenido completo (copies, arte, hashtags) se agrega como nota en cada ítem."
        )

        configured = _monday_configured()

        if not configured:
            st.warning("Monday.com no está configurado. Sigue estos pasos:")
            st.markdown("""
**Paso 1 — Obtén tu API key:**
- Ve a `monday.com → tu avatar → Developers → My Access Tokens`
- Copia el token personal

**Paso 2 — Obtén el ID de tu tablero:**
- Abre el tablero en Monday.com
- La URL tiene el formato: `monday.com/boards/**1234567890**`
- Ese número es el `MONDAY_BOARD_ID`

**Paso 3 — Agrega al archivo `.env`:**
""")
            st.code("MONDAY_API_KEY=eyJhbGciOiJIUzI1...\nMONDAY_BOARD_ID=1234567890", language="bash")
            env_path = ROOT / '.env'
            st.caption(f"Archivo: `{env_path}`")

            with st.expander("Editar .env directamente aquí"):
                if env_path.exists():
                    current_env = env_path.read_text(encoding='utf-8')
                else:
                    current_env = "ANTHROPIC_API_KEY=\n"
                new_monday_key  = st.text_input("MONDAY_API_KEY", type="password",
                                                key="monday_key_input",
                                                placeholder="eyJhbGciOiJIUzI1...")
                new_monday_board = st.text_input("MONDAY_BOARD_ID",
                                                 key="monday_board_input",
                                                 placeholder="1234567890")
                if st.button("Guardar credenciales Monday.com", key="btn_save_monday_creds"):
                    if new_monday_key and new_monday_board:
                        lines = current_env.splitlines()
                        new_lines, added_key, added_board = [], False, False
                        for line in lines:
                            if line.startswith('MONDAY_API_KEY='):
                                new_lines.append(f"MONDAY_API_KEY={new_monday_key}")
                                added_key = True
                            elif line.startswith('MONDAY_BOARD_ID='):
                                new_lines.append(f"MONDAY_BOARD_ID={new_monday_board}")
                                added_board = True
                            else:
                                new_lines.append(line)
                        if not added_key:
                            new_lines.append(f"MONDAY_API_KEY={new_monday_key}")
                        if not added_board:
                            new_lines.append(f"MONDAY_BOARD_ID={new_monday_board}")
                        env_path.write_text('\n'.join(new_lines) + '\n', encoding='utf-8')
                        # Reset env cache so next call picks up new values
                        os.environ.pop('MONDAY_API_KEY', None)
                        os.environ.pop('MONDAY_BOARD_ID', None)
                        st.success("Credenciales guardadas. Recarga la página para activar.")
                        st.rerun()
                    else:
                        st.error("Completa ambos campos.")
        else:
            api_key  = _monday_api_key()
            board_id = _monday_board_id()

            # Board info preview
            board_info_key = f"monday_board_info_{board_id}"
            if board_info_key not in st.session_state:
                try:
                    bname, bcols, _ = _monday_get_board_info(api_key, board_id)
                    st.session_state[board_info_key] = {'name': bname, 'cols': bcols}
                except Exception as e:
                    st.session_state[board_info_key] = {'error': str(e)}

            bi = st.session_state.get(board_info_key, {})
            if 'error' in bi:
                st.error(f"Error al conectar con Monday.com: {bi['error']}")
            else:
                bname = bi.get('name', '—')
                bcols = bi.get('cols', [])
                col_map = _monday_match_columns(bcols)
                st.success(f"Conectado a tablero: **{bname}**")

                col_info2, col_map_info = st.columns([1, 1])
                with col_info2:
                    st.metric("Items a crear", len(st.session_state['parrilla_df']))
                with col_map_info:
                    mapped = ', '.join(col_map.keys()) or 'ninguno'
                    st.caption(f"Columnas detectadas: {mapped} · Resto se agrega como nota")

                with st.expander("🔍 Diagnóstico de columnas del tablero", expanded=False):
                    st.markdown("**Todas las columnas en Monday.com:**")
                    for c in bcols:
                        matched_as = next(
                            (k for k, v in col_map.items() if v['id'] == c['id']), None
                        )
                        tag = f" → **mapeada como `{matched_as}`**" if matched_as else ""
                        st.markdown(f"- `{c['id']}` · **{c['title']}** · tipo: `{c['type']}`{tag}")
                    estado_info = col_map.get('estado')
                    if estado_info:
                        st.info(
                            f"Estado se actualizará en columna **'{estado_info['id']}'** "
                            f"(tipo: `{estado_info['type']}`)"
                        )
                    else:
                        st.warning("No se detectó columna Estado — los estados NO se sincronizarán.")

            st.markdown("---")

            # Reset button — always visible so user can re-sync after Monday board changes
            from database import get_parrilla_posts as _gpp, clear_monday_item_ids
            db_posts_mon = _gpp(meta.get('marca', ''), año, mes)
            has_ids = any(p.get('monday_item_id') for p in db_posts_mon)

            col_sync_btn, col_reset = st.columns([3, 1])
            with col_reset:
                if st.button(
                    "🗑️ Limpiar IDs",
                    use_container_width=True,
                    key="btn_monday_reset",
                    disabled=not has_ids,
                    help="Borra los IDs de Monday guardados localmente. "
                         "Úsalo cuando hayas borrado ítems o columnas en Monday "
                         "y quieras volver a sincronizar desde cero."
                ):
                    clear_monday_item_ids(meta.get('marca', ''), año, mes)
                    st.session_state.pop(f"monday_sync_{marca_key}_{año}_{mes}", None)
                    st.success("IDs limpiados. Ahora sincroniza de nuevo.")
                    st.rerun()

            # Sync button
            sync_key = f"monday_sync_{marca_key}_{año}_{mes}"
            with col_sync_btn:
                if st.button("🔄  Sincronizar parrilla con Monday.com",
                             type="primary", use_container_width=True,
                             key="btn_monday_sync"):
                    with st.spinner("Creando ítems en Monday.com…"):
                        try:
                            result = _monday_sync_parrilla(
                                st.session_state['parrilla_df'],
                                meta.get('marca', ''), año, mes, api_key, board_id,
                            )
                            st.session_state[sync_key] = result
                        except Exception as e:
                            st.error(f"Error durante la sincronización: {e}")

            # Show results
            if sync_key in st.session_state:
                res = st.session_state[sync_key]
                c1, c2 = st.columns(2)
                with c1:
                    st.metric("✅ Ítems creados", res.get('created', 0))
                with c2:
                    st.metric("⏭️ Ya sincronizados", res.get('skipped', 0))

                items = res.get('results', [])
                if items:
                    st.markdown("**Ítems en Monday.com:**")
                    for item in items:
                        prefix = "🆕" if item.get('new') else "✅"
                        item_id = item.get('item_id', '')
                        link = f"https://monday.com/boards/{board_id}/pulses/{item_id}"
                        st.markdown(
                            f"{prefix} {item.get('fecha','')} ({item.get('tipo_dia','')}) — "
                            f"[Ver en Monday.com]({link})"
                        )

    # ═══ TAB 4: PROMPTS DE IMAGEN ════════════════════════════════════════════════
    if not _is_visita:
     with tab4:

        # ── Editor de imagen propia ────────────────────────────────────────────
        with st.expander("📤 Editar imagen propia", expanded=False):
            st.caption(
                "Sube una imagen tuya y aplica ediciones con IA o agrega el logo de la marca."
            )

            _up_file = st.file_uploader(
                "Sube tu imagen",
                type=["png", "jpg", "jpeg", "webp"],
                key="img_upload_file",
            )

            _up_cache = "img_upload_current"
            _up_edits = "img_upload_edits"

            if _up_file is not None:
                _raw_bytes = _up_file.read()
                from PIL import Image as _PILImage
                _pil = _PILImage.open(io.BytesIO(_raw_bytes)).convert('RGB')
                _buf = io.BytesIO()
                _pil.save(_buf, format='PNG')
                _orig_bytes = _buf.getvalue()
                # Reset history when a new file is uploaded
                if st.session_state.get("img_upload_last_name") != _up_file.name:
                    st.session_state["img_upload_last_name"] = _up_file.name
                    st.session_state[_up_cache] = _orig_bytes
                    st.session_state[_up_edits] = []

            if _up_cache in st.session_state:
                _up_hist    = st.session_state.setdefault(_up_edits, [])
                _up_current = _up_hist[-1]['bytes'] if _up_hist else st.session_state[_up_cache]
                _up_ver     = len(_up_hist) + 1

                _col_up_img, _col_up_ctrl = st.columns([1, 1], gap="large")

                with _col_up_img:
                    st.image(_up_current, use_container_width=True)
                    if _up_hist:
                        st.caption(f"Versión {_up_ver} · {len(_up_hist)} edición(es)")

                    st.download_button(
                        f"📥 Descargar v{_up_ver}",
                        data=_up_current,
                        file_name=f"imagen_editada_v{_up_ver}.png",
                        mime="image/png",
                        use_container_width=True,
                        key="btn_dl_upload",
                    )
                    _logo_res_up = _logo_adder_ui("up_editor_main", _up_current, brand)
                    if _logo_res_up is not None:
                        if _up_hist and _up_hist[-1].get('instruction') == '[logo agregado]':
                            _up_hist[-1] = {'instruction': '[logo agregado]', 'bytes': _logo_res_up}
                        else:
                            _up_hist.append({'instruction': '[logo agregado]', 'bytes': _logo_res_up})
                        st.rerun()

                with _col_up_ctrl:
                    st.markdown("**Editar con IA**")
                    _up_instr = st.text_area(
                        "Instrucción de edición",
                        placeholder=(
                            "Ejemplos de lo que puedes pedir:\n"
                            "• Cambia el título a 'Seguridad Pública Inteligente'\n"
                            "• Agrega el texto 'Kabat One' en la parte inferior\n"
                            "• Reemplaza el texto del botón por 'Contáctanos'\n"
                            "• Cambia el fondo a azul oscuro\n"
                            "• Agrega iluminación nocturna con tonos neón"
                        ),
                        height=140,
                        key="img_upload_instr",
                    )

                    _col_up_apply, _col_up_undo = st.columns([3, 1])
                    with _col_up_apply:
                        _up_apply = st.button(
                            "🪄 Aplicar edición",
                            type="primary",
                            use_container_width=True,
                            disabled=not _up_instr.strip(),
                            key="btn_up_apply",
                        )
                    with _col_up_undo:
                        if st.button(
                            "↩️",
                            use_container_width=True,
                            disabled=not _up_hist,
                            key="btn_up_undo",
                            help="Deshacer última edición",
                        ):
                            _up_hist.pop()
                            st.rerun()

                    if _up_hist:
                        st.markdown("**Historial:**")
                        for _ui, _uh in enumerate(_up_hist):
                            st.caption(f"v{_ui+2}: {_uh['instruction'][:70]}")

                    if _up_apply and _up_instr.strip():
                        _up_err_ph = st.empty()
                        with st.spinner("Aplicando edición… (15-30 segundos)"):
                            try:
                                _up_edited = _edit_imagen_gemini(_up_current, _up_instr.strip())
                                _up_hist.append({
                                    'instruction': _up_instr.strip(),
                                    'bytes': _up_edited,
                                })
                                st.rerun()
                            except Exception as _ue:
                                _up_err_ph.error(f"Error al editar: {_ue}")

        # ── Generador de Carrusel ──────────────────────────────────────────────
        with st.expander("🎠 Generador de Carrusel", expanded=False):
            st.caption(
                "La IA diseña la estructura narrativa del carrusel y Imagen 4 genera "
                "cada slide con coherencia visual. Descarga slides individuales o en ZIP."
            )

            _car_opciones = []
            for _, _crow in df.iterrows():
                _cf = str(_crow.get('Fecha', ''))
                _ct = str(_crow.get('Tema', ''))[:50]
                _ctp = str(_crow.get('Tipo', ''))
                _car_opciones.append(f"{_cf} · {_ct}" + (" ⭐" if _ctp == 'especial' else ""))

            if not _car_opciones:
                st.info("No hay publicaciones en la parrilla.")
            else:
                _car_sel_lbl = st.selectbox("Publicación base", _car_opciones, key="car_pub_sel")
                _car_sel_row = df.iloc[_car_opciones.index(_car_sel_lbl)].to_dict()

                _cc1, _cc2, _cc3 = st.columns(3)
                with _cc1:
                    _car_n = st.selectbox("Slides", [3, 4, 5, 6, 7, 8], index=2, key="car_n_slides")
                with _cc2:
                    _car_style_opts = {
                        '🃏 Cards': 'cards', '🏙️ Isométrico': 'isometric', '📊 Infografía': 'infographic',
                    }
                    _car_style_sel = st.selectbox("Estilo", list(_car_style_opts.keys()), key="car_style_sel")
                    _car_style     = _car_style_opts[_car_style_sel]
                with _cc3:
                    _car_q_opts = {"⚡ Rápida": "fast", "⭐ Estándar": "standard", "💎 Ultra": "ultra"}
                    _car_q_sel  = st.selectbox("Calidad", list(_car_q_opts.keys()), index=1, key="car_quality_sel")
                    _car_q      = _car_q_opts[_car_q_sel]

                _car_struct_key = f"car_struct_{meta.get('marca','')}_{_car_sel_lbl}_{_car_n}_{_car_style}"
                _car_imgs_key   = f"car_imgs_{_car_struct_key}_{_car_q}"

                _car_prov_lbl = "Gemini" if _ai_provider() == 'gemini' else "Claude"
                _cs1, _cs2 = st.columns([3, 1])
                with _cs1:
                    if st.button(
                        f"✨ Diseñar estructura con {_car_prov_lbl}",
                        type="primary", use_container_width=True, key="btn_car_struct",
                    ):
                        st.session_state.pop(_car_struct_key, None)
                        st.session_state.pop(_car_imgs_key, None)
                        with st.spinner(f"Diseñando {_car_n} slides…"):
                            try:
                                _cp = _build_carousel_slides_prompt(_car_sel_row, brand, _car_n, _car_style)
                                _cr = _call_claude_json(_cp)
                                st.session_state[_car_struct_key] = _parse_json_obj(_cr)
                            except Exception as _ce:
                                st.error(f"Error: {_ce}")
                with _cs2:
                    if st.button("🗑️ Reiniciar", use_container_width=True, key="btn_car_reset"):
                        st.session_state.pop(_car_struct_key, None)
                        st.session_state.pop(_car_imgs_key, None)
                        st.rerun()

                if _car_struct_key in st.session_state:
                    _car_data  = st.session_state[_car_struct_key]
                    _car_slides = _car_data.get('slides', [])
                    st.markdown(f"**{_car_data.get('titulo_carrusel', '')}**")

                    _tipo_icon = {'portada': '🎯', 'contenido': '📌', 'cta': '📣'}
                    for _sl in _car_slides:
                        _ico = _tipo_icon.get(_sl.get('tipo',''), '📌')
                        with st.expander(f"{_ico} Slide {_sl['numero']}: {_sl.get('titulo_slide','')}", expanded=False):
                            _sl['escena'] = st.text_area(
                                "Escena (editable antes de generar)",
                                value=_sl.get('escena', ''),
                                height=90,
                                key=f"car_escena_{_sl['numero']}",
                            )

                    _car_time_map = {"fast": 10, "standard": 25, "ultra": 45}
                    _total_min    = (_car_time_map.get(_car_q, 25) * _car_n) // 60 + 1
                    st.caption(f"Tiempo estimado: ~{_total_min} min · Formato 1:1 (carrusel Instagram/LinkedIn)")

                    if st.button(
                        f"🎨 Generar {_car_n} slides",
                        type="primary", use_container_width=True, key="btn_car_gen",
                    ):
                        st.session_state.pop(_car_imgs_key, None)
                        _master    = _MASTER_PROMPTS.get(_car_style, '')
                        _car_imgs  = []
                        _cprog     = st.progress(0, text="Iniciando…")
                        _cerr_ph   = st.empty()

                        for _ci, _sl in enumerate(_car_slides):
                            _cprog.progress(
                                _ci / _car_n,
                                text=f"Slide {_sl['numero']}/{_car_n} — {_sl.get('titulo_slide','')[:45]}…"
                            )
                            _fp = f"{_sl['escena']}\n\n{_master}" if _master else _sl['escena']
                            try:
                                _sb = _generate_imagen(_fp, '1:1', _car_q)
                                _sb = _add_slide_text(
                                    _sb,
                                    _sl.get('titulo_slide', ''),
                                    _sl.get('subtitulo', ''),
                                    _sl['numero'],
                                    _car_n,
                                    brand_id=brand.get('brand_id', ''),
                                )
                                _car_imgs.append({
                                    'numero': _sl['numero'], 'tipo': _sl.get('tipo',''),
                                    'titulo': _sl.get('titulo_slide',''), 'bytes': _sb,
                                })
                            except Exception as _ge:
                                _cerr_ph.warning(f"Slide {_sl['numero']} falló: {_ge}")
                                _car_imgs.append({
                                    'numero': _sl['numero'], 'tipo': _sl.get('tipo',''),
                                    'titulo': _sl.get('titulo_slide',''), 'bytes': None,
                                })

                        _ok_count = len([s for s in _car_imgs if s['bytes']])
                        _cprog.progress(1.0, text=f"✅ {_ok_count}/{_car_n} slides generados")
                        st.session_state[_car_imgs_key] = _car_imgs
                        st.rerun()

                    if _car_imgs_key in st.session_state:
                        _car_imgs = st.session_state[_car_imgs_key]
                        _car_edit_key = _car_imgs_key + '_edit_slide'
                        st.markdown("---")
                        _gc = st.columns(min(3, len(_car_imgs)))
                        for _gi, _gsl in enumerate(_car_imgs):
                            with _gc[_gi % 3]:
                                _ico = _tipo_icon.get(_gsl.get('tipo',''), '📌')
                                if _gsl['bytes']:
                                    st.image(_gsl['bytes'], use_container_width=True)
                                    st.caption(f"{_ico} {_gsl['titulo'][:35]}")
                                    _mrc_z = meta.get('marca','').replace(' ','_')
                                    st.download_button(
                                        f"📥 Slide {_gsl['numero']}",
                                        data=_gsl['bytes'],
                                        file_name=f"carrusel_{_mrc_z}_slide_{_gsl['numero']:02d}.png",
                                        mime="image/png",
                                        use_container_width=True,
                                        key=f"btn_dl_sl_{_gsl['numero']}",
                                    )
                                    _logo_res_sl = _logo_adder_ui(
                                        f"car_{_car_imgs_key}_sl{_gsl['numero']}",
                                        _gsl['bytes'], brand,
                                    )
                                    if _logo_res_sl is not None:
                                        for _upd in st.session_state[_car_imgs_key]:
                                            if _upd['numero'] == _gsl['numero']:
                                                _upd['bytes'] = _logo_res_sl
                                                break
                                        st.rerun()
                                    _edit_open = st.session_state.get(_car_edit_key) == _gsl['numero']
                                    _btn_lbl = "✏️ Cerrar editor" if _edit_open else "✏️ Editar slide"
                                    if st.button(_btn_lbl, key=f"btn_edit_sl_{_gsl['numero']}", use_container_width=True):
                                        if _edit_open:
                                            st.session_state.pop(_car_edit_key, None)
                                        else:
                                            st.session_state[_car_edit_key] = _gsl['numero']
                                        st.rerun()
                                    if _edit_open:
                                        _edit_inst = st.text_area(
                                            "Instrucción de edición",
                                            placeholder=(
                                                "Ej: Cambia el título a 'Innovación'\n"
                                                "Agrega texto 'Kabat One' abajo\n"
                                                "Cambia fondo a azul oscuro"
                                            ),
                                            height=100,
                                            key=f"ta_edit_sl_{_gsl['numero']}",
                                            label_visibility="collapsed",
                                        )
                                        if st.button("🪄 Aplicar edición", key=f"btn_apply_sl_{_gsl['numero']}", type="primary", use_container_width=True):
                                            if _edit_inst.strip():
                                                with st.spinner("Editando con IA…"):
                                                    try:
                                                        _new_bytes = _edit_imagen_gemini(_gsl['bytes'], _edit_inst.strip())
                                                        for _upd in st.session_state[_car_imgs_key]:
                                                            if _upd['numero'] == _gsl['numero']:
                                                                _upd['bytes'] = _new_bytes
                                                                break
                                                        st.session_state.pop(_car_edit_key, None)
                                                        st.rerun()
                                                    except Exception as _ee:
                                                        st.error(f"Error al editar: {_ee}")
                                            else:
                                                st.warning("Escribe una instrucción.")
                                else:
                                    st.error(f"Slide {_gsl['numero']} falló")

                        _slides_ok = [s for s in _car_imgs if s['bytes']]
                        if len(_slides_ok) > 1:
                            import zipfile
                            _zbuf = io.BytesIO()
                            _mrc_z = meta.get('marca','').replace(' ','_')
                            with zipfile.ZipFile(_zbuf, 'w', zipfile.ZIP_DEFLATED) as _zf:
                                for _zsl in _slides_ok:
                                    _zf.writestr(
                                        f"carrusel_{_mrc_z}_slide_{_zsl['numero']:02d}_{_zsl['tipo']}.png",
                                        _zsl['bytes'],
                                    )
                            st.download_button(
                                f"📦 Descargar todos ({len(_slides_ok)} slides) .zip",
                                data=_zbuf.getvalue(),
                                file_name=f"carrusel_{_mrc_z}_{_car_sel_row.get('Fecha','')}.zip",
                                mime="application/zip",
                                use_container_width=True,
                                key="btn_dl_car_zip",
                            )

        st.markdown("---")
        st.markdown("#### 🎨 Generador de Prompts de Imagen")
        st.caption(
            "Selecciona una publicación y genera prompts profesionales listos para pegar "
            "en **DALL-E 3 (ChatGPT)** y **Gemini Imagen** (Google)."
        )

        _prov_img = _ai_provider()
        _key_img_ok = _get_gemini_key() if _prov_img == 'gemini' else _get_api_key()
        if not _key_img_ok:
            _needed = 'GEMINI_API_KEY' if _prov_img == 'gemini' else 'ANTHROPIC_API_KEY'
            st.warning(f"Configura `{_needed}` en `.env` para usar este módulo.")
        else:
            # Selector de publicación
            opciones = []
            for _, row in df.iterrows():
                fecha = str(row.get('Fecha', ''))
                tema  = str(row.get('Tema', ''))[:50]
                tipo  = str(row.get('Tipo', ''))
                lbl   = f"{fecha} · {tema}" + (" ⭐" if tipo == 'especial' else "")
                opciones.append(lbl)

            if not opciones:
                st.info("No hay publicaciones en la parrilla.")
            else:
                sel_lbl = st.selectbox("Publicación", opciones, key="img_prompt_sel")
                sel_idx = opciones.index(sel_lbl)
                sel_row = df.iloc[sel_idx].to_dict()

                col_red, _ = st.columns([2, 3])
                with col_red:
                    red_img = st.radio(
                        "Formato destino",
                        ["LinkedIn (1200×628)", "Instagram / Facebook (1080×1080)", "Ambos"],
                        horizontal=True,
                        key="img_prompt_red",
                    )

                st.markdown("**Estilo visual**")
                _style_cols = st.columns(len(_STYLE_OPTIONS))
                _style_sel_key = "img_style_sel"
                if _style_sel_key not in st.session_state:
                    st.session_state[_style_sel_key] = '🏙️ Isométrico'
                for _i, (_slabel, _sval) in enumerate(_STYLE_OPTIONS.items()):
                    with _style_cols[_i]:
                        _is_active = st.session_state[_style_sel_key] == _slabel
                        if st.button(
                            _slabel,
                            key=f"style_btn_{_sval}",
                            use_container_width=True,
                            type="primary" if _is_active else "secondary",
                        ):
                            st.session_state[_style_sel_key] = _slabel
                            st.rerun()

                _style_sel   = st.session_state.get(_style_sel_key, '🏙️ Isométrico')
                _style_value = _STYLE_OPTIONS.get(_style_sel, 'libre')

                _style_desc = {
                    'isometric':   '🏙️ Diorama miniatura de ciudad inteligente con plataforma flotante y luces azules',
                    'cards':       '🃏 Dispositivo hero con entorno real fotorrealista y paneles de datos flotantes',
                    'infographic': '📊 Ciudad digital twin con visualización de datos operativos y glassmorphism',
                    'libre':       '✏️ El asistente define el estilo completo basado en el brief de marca',
                }
                st.caption(_style_desc.get(_style_value, ''))

                _img_error_ph = st.empty()

                if st.button("✨  Generar Prompts de Imagen", type="primary",
                             use_container_width=True, key="btn_gen_img_prompt"):
                    cache_key = f"img_prompts_{meta.get('marca','')}_{sel_lbl}_{red_img}_{_style_value}"
                    st.session_state.pop(cache_key, None)
                    _img_error_ph.empty()

                    prompt_req = _build_image_prompt_request(sel_row, brand, red_img, _style_value)
                    _ai_lbl_img = "Gemini" if _ai_provider() == 'gemini' else "Claude"
                    _spin_ph = st.empty()
                    with _spin_ph:
                        with st.spinner(f"{_ai_lbl_img} está creando los prompts… (10-20 seg)"):
                            try:
                                raw = _call_claude_json(prompt_req)
                                result = _parse_json_obj(raw)
                                if not result:
                                    result = {'error': 'No se pudo parsear la respuesta', 'raw': raw}
                                st.session_state[cache_key] = result
                            except Exception as e:
                                err_msg = str(e)
                                if 'credit' in err_msg.lower() or 'balance' in err_msg.lower():
                                    _img_error_ph.error(
                                        "❌ Saldo insuficiente en Anthropic. "
                                        "Cambia a Gemini (gratis) en el sidebar, o agrega créditos en "
                                        "[console.anthropic.com](https://console.anthropic.com)."
                                    )
                                else:
                                    _img_error_ph.error(f"Error al generar: {err_msg}")
                    _spin_ph.empty()

                # Mostrar resultado
                cache_key = f"img_prompts_{meta.get('marca','')}_{sel_lbl}_{red_img}_{_style_value}"
                if cache_key in st.session_state:
                    result = st.session_state[cache_key]
                    if 'error' in result:
                        st.error(result['error'])
                        if 'raw' in result:
                            with st.expander("Respuesta cruda"):
                                st.text(result['raw'])
                    else:
                        desc = result.get('descripcion_corta', '')
                        _style_badge = '' if _style_value == 'libre' else f" · Estilo: **{_style_sel}**"
                        if desc:
                            st.markdown(f"**Imagen:** {desc}{_style_badge}")
                        st.markdown("---")

                        col_dalle, col_gemini = st.columns(2)

                        with col_dalle:
                            st.markdown("### 🤖 DALL-E 3 · ChatGPT")
                            dalle = result.get('dalle3', {})
                            prompt_text = dalle.get('prompt', '')
                            _full_dalle = f"{prompt_text}\n\n{_MASTER_PROMPTS[_style_value]}" if _style_value in _MASTER_PROMPTS else prompt_text
                            st.text_area(
                                "Prompt completo (escena + estilo master)",
                                value=_full_dalle,
                                height=260,
                                key="dalle_prompt_area",
                            )
                            notas = dalle.get('notas', '')
                            if notas:
                                st.caption(f"💡 {notas}")
                            if _style_value in _MASTER_PROMPTS:
                                st.caption(f"✅ Master **{_style_sel}** incluido automáticamente")

                        with col_gemini:
                            st.markdown("### 🌐 Gemini Imagen · Google")
                            gemini = result.get('gemini', {})
                            prompt_text_g = gemini.get('prompt', '')
                            _full_gemini = f"{prompt_text_g}\n\n{_MASTER_PROMPTS[_style_value]}" if _style_value in _MASTER_PROMPTS else prompt_text_g
                            st.text_area(
                                "Prompt completo (escena + estilo master)",
                                value=_full_gemini,
                                height=260,
                                key="gemini_prompt_area",
                            )
                            notas_g = gemini.get('notas', '')
                            if notas_g:
                                st.caption(f"💡 {notas_g}")
                            if _style_value in _MASTER_PROMPTS:
                                st.caption(f"✅ Master **{_style_sel}** incluido automáticamente")

                        # Guía de uso
                        with st.expander("📖 ¿Cómo usar estos prompts?"):
                            st.markdown("""
**ChatGPT / DALL-E 3:**
1. Abre [chatgpt.com](https://chatgpt.com) y crea una nueva conversación
2. Haz clic en el ícono de imagen (🖼️) o escribe el prompt directamente
3. Si quieres variaciones, escribe: *"genera 4 variaciones de este prompt"*

**Gemini Imagen (Google):**
1. Ve a [gemini.google.com](https://gemini.google.com)
2. Pega el prompt en el chat
3. Gemini generará la imagen automáticamente

**Consejo:** Después de generar, pide ajustes con frases como:
- *"Hazla más minimalista"*
- *"Cambia el fondo a azul oscuro"*
- *"Versión sin personas, solo tecnología"*
""")

                        # ── Generación directa con Imagen 4 ───────────────
                        st.markdown("---")
                        st.markdown("### 🖼️ Generar imagen aquí con Imagen 4")
                        st.caption(
                            "Usa el prompt de Gemini como base para generar la imagen directamente. "
                            "Descárgala y úsala en tu diseño."
                        )

                        _scene_desc = result.get('gemini', {}).get('prompt', '')
                        _master     = _MASTER_PROMPTS.get(_style_value, '')
                        _prompt_gen = f"{_scene_desc}\n\n{_master}" if _master else _scene_desc
                        _img4_cache = f"img4_{cache_key}"

                        # Aspect ratio selector si es "Ambos"
                        if 'LinkedIn' in red_img:
                            _aspect4 = '16:9'
                            _aspect_lbl = 'LinkedIn 16:9'
                        elif 'Instagram' in red_img or 'Facebook' in red_img:
                            _aspect4 = '1:1'
                            _aspect_lbl = 'Instagram/Facebook 1:1'
                        else:
                            _col_asp, _ = st.columns([2, 3])
                            with _col_asp:
                                _asp_sel = st.radio(
                                    "Formato a generar",
                                    ["LinkedIn (16:9)", "Instagram (1:1)"],
                                    horizontal=True,
                                    key="img4_aspect_sel",
                                )
                            _aspect4    = '16:9' if 'LinkedIn' in _asp_sel else '1:1'
                            _aspect_lbl = _asp_sel

                        _col_qual, _col_info = st.columns([3, 2])
                        with _col_qual:
                            _quality_sel = st.radio(
                                "Motor de imagen",
                                ["⚡ Rápida", "⭐ Estándar", "💎 Ultra", "🍌 Nano Banana"],
                                index=1,
                                horizontal=True,
                                key="img4_quality",
                            )
                        _quality_map = {
                            "⚡ Rápida":      "fast",
                            "⭐ Estándar":    "standard",
                            "💎 Ultra":       "ultra",
                            "🍌 Nano Banana": "nano_banana",
                        }
                        _quality = _quality_map[_quality_sel]

                        with _col_info:
                            _time_est = {
                                "fast":        "~10 seg",
                                "standard":    "~20 seg",
                                "ultra":       "~40 seg",
                                "nano_banana": "~30 seg · con razonamiento",
                            }
                            st.caption(f"Formato: **{_aspect_lbl}** · {_time_est[_quality]}")

                        _gen_img = st.button(
                            "🎨  Generar imagen con Imagen 4",
                            type="primary",
                            use_container_width=True,
                            key="btn_gen_imagen4",
                            disabled=not _prompt_gen,
                        )

                        _img4_err = st.empty()

                        _edit_hist_key = f"img4_edits_{_img4_cache}"

                        if _gen_img:
                            st.session_state.pop(_img4_cache, None)
                            st.session_state.pop(_edit_hist_key, None)
                            _spin_lbl = {"fast": "10-15", "standard": "20-30", "ultra": "35-50", "nano_banana": "30-60"}
                            with st.spinner(f"Generando… ({_spin_lbl.get(_quality, '20-40')} segundos)"):
                                try:
                                    _img_bytes = _generate_imagen(_prompt_gen, _aspect4, _quality)
                                    st.session_state[_img4_cache] = {
                                        'bytes': _img_bytes,
                                        'aspect': _aspect_lbl,
                                    }
                                except Exception as _e:
                                    _img4_err.error(f"Error al generar imagen: {_e}")

                        if _img4_cache in st.session_state:
                            _saved      = st.session_state[_img4_cache]
                            _edit_hist  = st.session_state.setdefault(_edit_hist_key, [])
                            _cur_bytes  = _edit_hist[-1]['bytes'] if _edit_hist else _saved['bytes']
                            _ver_num    = len(_edit_hist) + 1

                            _fecha_dl = str(sel_row.get('Fecha', ''))
                            _tema_dl  = str(sel_row.get('Tema', ''))[:30].replace(' ', '_')
                            _fname_base = f"imagen_{meta.get('marca','').replace(' ','_')}_{_fecha_dl}_{_tema_dl}"

                            st.markdown("---")
                            _col_img_ed, _col_ctrl_ed = st.columns([1, 1], gap="large")

                            with _col_img_ed:
                                st.image(_cur_bytes, use_container_width=True)
                                if _edit_hist:
                                    st.caption(f"Versión {_ver_num} · {len(_edit_hist)} edición(es) aplicada(s)")
                                st.download_button(
                                    f"📥 Descargar v{_ver_num} (.png)",
                                    data=_cur_bytes,
                                    file_name=f"{_fname_base}_v{_ver_num}.png",
                                    mime="image/png",
                                    use_container_width=True,
                                    key="btn_dl_imagen4",
                                )
                                _logo_res_img4 = _logo_adder_ui(f"img4_{_img4_cache}", _cur_bytes, brand)
                                if _logo_res_img4 is not None:
                                    if _edit_hist and _edit_hist[-1].get('instruction') == '[logo agregado]':
                                        _edit_hist[-1] = {'instruction': '[logo agregado]', 'bytes': _logo_res_img4}
                                    else:
                                        _edit_hist.append({'instruction': '[logo agregado]', 'bytes': _logo_res_img4})
                                    st.rerun()

                            with _col_ctrl_ed:
                                st.markdown("#### ✏️ Editar imagen con IA")
                                st.caption(
                                    "Describe qué quieres cambiar. "
                                    "Cada edición genera una nueva versión."
                                )

                                _edit_instr = st.text_area(
                                    "Instrucción de edición",
                                    placeholder=(
                                        "Ejemplos de lo que puedes pedir:\n"
                                        "• Cambia el título a 'Innovación en Seguridad'\n"
                                        "• Agrega el texto 'Soluciones para el Sector Público'\n"
                                        "• Quita el texto y deja solo la imagen\n"
                                        "• Cambia el cielo a atardecer\n"
                                        "• Agrega más vegetación en primer plano\n"
                                        "• Hazla más nocturna con iluminación azul"
                                    ),
                                    height=150,
                                    key="img4_edit_instr",
                                )

                                _col_apply, _col_undo = st.columns([3, 1])
                                with _col_apply:
                                    _apply_edit = st.button(
                                        "🪄 Aplicar edición",
                                        type="primary",
                                        use_container_width=True,
                                        disabled=not _edit_instr.strip(),
                                        key="btn_apply_edit",
                                    )
                                with _col_undo:
                                    _can_undo = bool(_edit_hist)
                                    if st.button(
                                        "↩️",
                                        use_container_width=True,
                                        disabled=not _can_undo,
                                        key="btn_undo_edit",
                                        help="Deshacer última edición",
                                    ):
                                        _edit_hist.pop()
                                        st.rerun()

                                if _edit_hist:
                                    st.markdown("**Historial de ediciones:**")
                                    for _ei, _eh in enumerate(_edit_hist):
                                        st.caption(f"v{_ei+2}: {_eh['instruction'][:70]}")

                                if _apply_edit and _edit_instr.strip():
                                    _edit_err_ph = st.empty()
                                    with st.spinner("Aplicando edición… (15-30 segundos)"):
                                        try:
                                            _edited = _edit_imagen_gemini(_cur_bytes, _edit_instr.strip())
                                            _edit_hist.append({
                                                'instruction': _edit_instr.strip(),
                                                'bytes': _edited,
                                            })
                                            st.rerun()
                                        except Exception as _ee:
                                            _edit_err_ph.error(f"Error al editar: {_ee}")
